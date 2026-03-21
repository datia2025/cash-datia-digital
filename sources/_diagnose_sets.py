from openpyxl import load_workbook
import os

BASE_DIR = r"c:\Users\gutie\OneDrive\FERFILES\DATIA\PROPUESTAS\COLOMBIA\+CONSULTA\LiquidityDashboard\sources"
MASTER_FILE = os.path.join(BASE_DIR, "Master Account.xlsx")

def diagnose_sets():
    wb = load_workbook(MASTER_FILE, read_only=True, data_only=True)
    ws_accounts = wb['Accounts']
    ws_list = wb['List']
    group_termino = {str(row[5]): str(row[7]) for row in ws_list.iter_rows(min_row=2, values_only=True) if row and len(row) >= 8}
    
    # Audit Logic Sets
    audit_ac, audit_pc = set(), set()
    for row in ws_accounts.iter_rows(min_row=2, values_only=True):
        if not row or not row[1]: continue
        c, g = str(row[1]).strip(), str(row[13])
        t = group_termino.get(g, "")
        if c.startswith('1') and (t == 'Corto Termino' or g in ('11','12','13','14','18','19')): audit_ac.add(c)
        if c.startswith('2') and (t == 'Corto Termino' or g in ('22','23','24')): audit_pc.add(c)
    
    print(f"Audit AC: {len(audit_ac)} accounts")
    print(f"Audit PC: {len(audit_pc)} accounts")

    # Re-running the logic that should be in calculate_indicators.py
    # (Checking if 'termino' is correctly parsed)
    
if __name__ == "__main__":
    diagnose_sets()
