from openpyxl import load_workbook
import csv
import os
from collections import defaultdict
from datetime import datetime

# Paths
BASE_DIR = r"c:\Users\gutie\OneDrive\FERFILES\DATIA\PROPUESTAS\COLOMBIA\+CONSULTA\LiquidityDashboard\sources"
MASTER_FILE = os.path.join(BASE_DIR, "Master Account.xlsx")
MOV_FILES = [os.path.join(BASE_DIR, f"Mov {y}.csv") for y in [2023, 2024, 2025]]
ORIGINAL_CSV = os.path.join(BASE_DIR, "LIQUIDEZ", "Razón Corriente.csv")

def audit_full_ytd():
    # 1. Load Ground Truth
    ground_truth = {}
    if os.path.exists(ORIGINAL_CSV):
        with open(ORIGINAL_CSV, "r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                y = int(row["Año Texto"])
                p = int(row["Period"])
                val = float(row["Times"])
                ground_truth[(y, p)] = val

    # 2. Load Classifications
    wb = load_workbook(MASTER_FILE, read_only=True, data_only=True)
    ws_accounts = wb['Accounts']
    ws_list = wb['List']
    
    group_termino = {}
    for row in ws_list.iter_rows(min_row=2, values_only=True):
        if row and len(row) >= 8:
            group_termino[str(row[5])] = str(row[7])

    ac_set = set()
    pc_set = set()
    
    for row in ws_accounts.iter_rows(min_row=2, values_only=True):
        if not row or not row[1]: continue
        codigo = str(row[1]).strip()
        clase = str(row[4])
        grupo = str(row[13])
        termino = group_termino.get(grupo, "")
        
        # Activo Corriente = 1 && Corto Termino
        if codigo.startswith('1') and (termino == 'Corto Termino' or grupo in ('11', '12', '13', '14', '18', '19')):
            ac_set.add(codigo)
        # Pasivo Corriente = 2 && Corto Termino
        if codigo.startswith('2') and (termino == 'Corto Termino' or grupo in ('22', '23', '24')):
            pc_set.add(codigo)
    
    # 3. Process all movements by year and month
    movements = defaultdict(lambda: defaultdict(float))
    for mov_path in MOV_FILES:
        if not os.path.exists(mov_path): continue
        with open(mov_path, 'r', encoding='utf-8', errors='replace') as f:
            reader = csv.reader(f)
            for row in reader:
                if not row or len(row) < 12 or not row[0].isdigit(): continue
                try:
                    dt = datetime.strptime(row[1].strip(), '%d/%m/%Y')
                    y, m = dt.year, dt.month
                    codigo = row[2].strip()
                    deb = float(row[10].replace(',', '')) if row[10] else 0.0
                    cre = float(row[11].replace(',', '')) if row[11] else 0.0
                    movements[(y, m)][codigo] += (deb - cre)
                except: continue

    # 4. Comparative Report
    print(f"{'Periodo':<10} | {'Calculated':<12} | {'Original':<12} | {'Delta':<10}")
    print("-" * 55)
    
    matches = 0
    total = 0

    for y in [2023, 2024, 2025]:
        # Reset YTD for current components at start of year (ALLSELECTED 'Calendario' behavior)
        ytd_ac = 0
        ytd_pc_neto = 0
        
        for m in range(1, 13):
            month_moves = movements.get((y, m), {})
            
            # Cumulative within the year
            ytd_ac += sum(month_moves.get(c, 0) for c in ac_set)
            ytd_pc_neto += sum(month_moves.get(c, 0) for c in pc_set)
            
            # PC in DAX = Neto * -1
            calc_val = ytd_ac / (ytd_pc_neto * -1) if ytd_pc_neto != 0 else 0
            
            orig_val = ground_truth.get((y, m), None)
            
            if orig_val is not None:
                total += 1
                delta = abs(calc_val - orig_val)
                if delta < 0.0001:
                    matches += 1
                    status = "OK"
                else:
                    status = f"{delta:.6f}"
                
                print(f"{y}-P{m:<2} | {calc_val:12.4f} | {orig_val:12.4f} | {status}")
            else:
                print(f"{y}-P{m:<2} | {calc_val:12.4f} | {'MISSING':<12} | -")

    print("-" * 55)
    print(f"Match Rate: {matches/total*100:.1f}% ({matches}/{total} periods)")

if __name__ == "__main__":
    audit_full_ytd()
