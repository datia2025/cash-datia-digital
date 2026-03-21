"""
PRUEBA ÁCIDA — Módulo de Actividad
MAS CONSULTA SAS - NIT 901271750

Reconstrucción independiente de 7 indicadores de Actividad
usando la lógica auditada y verificada al 100% contra Power BI.

Pipeline: Mov 20XX.csv + Master Account.xlsx → 7 CSVs

Lógica DAX replicada:
- Saldos acumulados históricos con ALL(Calendario) para promedios
- Exclusión de Documento 998 en Ventas, Compras y Costos (DSO/DIO/DPO)
- Rotación de Activos usa [Ingresos Operacionales] que INCLUYE Doc 998
- Calendario dinámico (días reales por mes)
- Promedio = (Saldo Inicial + Saldo Final) / 2
"""

import csv
import os
import calendar
from datetime import datetime
from collections import defaultdict
from openpyxl import load_workbook

# ============================================================
# CONFIG
# ============================================================
SOURCES_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'sources')
SOURCES_DIR = os.path.normpath(SOURCES_DIR)
OUTPUT_DIR = os.path.dirname(os.path.abspath(__file__))  # acid test folder
MASTER_FILE = os.path.join(SOURCES_DIR, "Master Account.xlsx")
MOV_FILES = [
    os.path.join(SOURCES_DIR, f"Mov {y}.csv") for y in [2021, 2022, 2023, 2024, 2025]
]
ANALYSIS_YEARS = [2023, 2024, 2025]


# ============================================================
# STEP 1: Load Master Account
# ============================================================
def load_account_classification():
    """Load account classification from Master Account.xlsx"""
    wb = load_workbook(MASTER_FILE, read_only=True, data_only=True)

    # Group → Termino mapping from List sheet
    ws_list = wb['List']
    group_termino = {}
    for row in ws_list.iter_rows(min_row=2, values_only=True):
        if row and len(row) >= 8:
            grupo = str(row[5]) if row[5] else None
            termino = str(row[7]) if row[7] else None
            if grupo and termino:
                group_termino[grupo] = termino

    # Accounts sheet
    ws_accounts = wb['Accounts']
    accounts = {}
    for row in ws_accounts.iter_rows(min_row=2, values_only=True):
        if not row or not row[1]:
            continue
        codigo = str(row[1]).strip()
        clase = str(row[4]) if row[4] else ""
        grupo = str(row[13]) if row[13] else ""
        termino = group_termino.get(grupo, "")

        accounts[codigo] = {
            'clase': clase,
            'grupo': grupo,
            'termino': termino,
        }

    wb.close()
    print(f"  Loaded {len(accounts)} accounts")
    return accounts


# ============================================================
# STEP 2: Parse Mov CSV files
# ============================================================
def parse_mov_files():
    """Parse all Mov 20XX.csv files. Returns movements dict."""
    movements = defaultdict(list)
    total_lines = 0

    for filepath in MOV_FILES:
        if not os.path.exists(filepath):
            print(f"  Warning: {filepath} not found.")
            continue

        filename = os.path.basename(filepath)
        print(f"  Parsing {filename}...")

        with open(filepath, 'r', encoding='utf-8', errors='replace') as f:
            reader = csv.reader(f)
            current_comprobante = ""
            file_lines = 0

            for row in reader:
                if not row:
                    continue

                first_val = row[0].strip()
                if first_val.startswith('Comprobante:'):
                    current_comprobante = first_val.replace('Comprobante:', '').strip()
                    continue

                if not first_val or not first_val[0].isdigit():
                    continue
                if len(row) < 12:
                    continue

                fecha_str = row[1].strip()
                codigo = row[2].strip()
                if not fecha_str or not codigo:
                    continue

                descripcion = row[7].strip() if len(row) > 7 else ""
                detalle = row[8].strip() if len(row) > 8 else ""

                is_closing = False
                if 'cierre anual' in descripcion.lower() or 'cierre anual' in detalle.lower():
                    is_closing = True
                if '998' in current_comprobante:
                    is_closing = True

                try:
                    fecha = datetime.strptime(fecha_str, '%d/%m/%Y')
                except ValueError:
                    continue

                year = fecha.year
                month = fecha.month

                try:
                    debito = float(row[10].replace(',', '')) if row[10] else 0.0
                    credito = float(row[11].replace(',', '')) if row[11] else 0.0
                except ValueError:
                    debito = 0.0
                    credito = 0.0

                key = (year, month, codigo)
                movements[key].append({
                    'debito': debito,
                    'credito': credito,
                    'is_closing': is_closing
                })
                total_lines += 1
                file_lines += 1

            print(f"    {file_lines} lines")

    print(f"  Total movements: {total_lines}")
    return movements


# ============================================================
# STEP 3: Build monthly balance snapshots
# ============================================================
def build_monthly_balances(movements):
    """
    Build cumulative balances (Full mode = ALL(Calendario)) 
    and monthly operational flows (excluding Doc 998).
    """
    sorted_keys = sorted(movements.keys())

    running_full = defaultdict(float)  # Full historical accumulation (includes 998)

    monthly_balances_full = {}
    monthly_flows_op = defaultdict(lambda: defaultdict(float))  # Excludes 998
    monthly_flows_all = defaultdict(lambda: defaultdict(float))  # Includes 998

    current_period = None

    for key in sorted_keys:
        year, month, codigo = key

        period = (year, month)
        if period != current_period:
            if current_period:
                monthly_balances_full[current_period] = dict(running_full)
            current_period = period

        for entry in movements[(year, month, codigo)]:
            flow = entry['debito'] - entry['credito']
            running_full[codigo] += flow
            monthly_flows_all[period][codigo] += flow  # ALL flows including 998

            if not entry['is_closing']:
                monthly_flows_op[period][codigo] += flow

    # Store final snapshot
    if current_period:
        monthly_balances_full[current_period] = dict(running_full)

    print(f"  Snapshots: {len(monthly_balances_full)} periods")
    return monthly_balances_full, monthly_flows_op, monthly_flows_all


# ============================================================
# STEP 4: Classify accounts for Activity indicators
# ============================================================
def classify_accounts(accounts):
    """
    Classify accounts into sets needed for Activity ratios.
    Aligned with DAX measures from metricas.md.
    """
    classifications = defaultdict(set)

    for codigo, data in accounts.items():
        grupo = data.get('grupo', '')
        first_digit = codigo[0] if codigo else '0'

        # Activos Totales (Clase 1) - for Rotación de Activos
        if first_digit == '1':
            classifications['activos_totales'].add(codigo)

        # CxC Cartera: DAX Accounts[Cuenta]=1305
        if grupo == '13' and codigo.startswith('1305'):
            classifications['cxc'].add(codigo)

        # Inventarios: DAX Accounts[Cuenta]=1435
        if grupo == '14' and codigo.startswith('1435'):
            classifications['inventarios'].add(codigo)

        # CxP Proveedores: DAX Accounts[Cuenta] in {2205, 2210, 2335}
        if first_digit == '2':
            if grupo == '22' and (codigo.startswith('2205') or codigo.startswith('2210')):
                classifications['cxp_proveedores'].add(codigo)
            elif grupo == '23' and codigo.startswith('2335'):
                classifications['cxp_proveedores'].add(codigo)

        # Ventas (Grupo 41): DAX Accounts[Grupo]=41 + Clase="Ingresos" + Doc<>998
        if first_digit == '4' and grupo == '41':
            classifications['ventas_41'].add(codigo)

        # Costos de Ventas (6135): DAX Accounts[Cuenta]=6135 + Doc<>998
        if first_digit in ('6', '7') and codigo.startswith('6135'):
            classifications['costos_ventas_6135'].add(codigo)

        # Compras: DAX set of accounts for DPO
        compras_grupos = (
            '5105', '5110', '5120', '5125', '5130', '5135', '5140', '5145',
            '5150', '5155', '5195', '5205', '5210', '5220', '5225', '5230',
            '5235', '5240', '5245', '5250', '5255', '5295'
        )
        if any(codigo.startswith(cg) for cg in compras_grupos):
            classifications['compras'].add(codigo)
        if codigo.startswith('6135'):
            classifications['compras'].add(codigo)

    # Print classification summary
    for key in sorted(classifications.keys()):
        print(f"    {key}: {len(classifications[key])} accounts")

    return classifications


# ============================================================
# STEP 5: Calculate Activity indicators
# ============================================================
def month_to_quarter(month):
    if month <= 3: return "1Q"
    if month <= 6: return "2Q"
    if month <= 9: return "3Q"
    return "4Q"


def safe_div(a, b):
    if b == 0 or b is None:
        return 0
    return a / b


def calculate_activity_indicators(monthly_balances_full, monthly_flows_op, classifications, monthly_flows_all=None):
    """
    Calculate 7 Activity indicators using audited logic.

    DAX Logic replicated:
    - Balances: ALL(Calendario) → Full historical accumulation from 2021
    - Flows (Ventas/Compras/Costos): Exclude Document 998
    - Average = (Saldo Inicial + Saldo Final) / 2
    - Saldo Inicial = Balance at (FechaInicioQ - 1) = previous period's end balance
    - Days = calendar.monthrange(year, month)[1] (matches COUNTROWS('Calendario'))
    """
    results = defaultdict(list)
    sorted_periods = sorted(monthly_balances_full.keys())

    for year in ANALYSIS_YEARS:
        for month in range(1, 13):
            key = (year, month)
            if key not in monthly_balances_full:
                continue

            q = month_to_quarter(month)
            bal = monthly_balances_full[key]
            flows = monthly_flows_op.get(key, {})
            flows_all = monthly_flows_all.get(key, {}) if monthly_flows_all else flows

            # Find previous period for average calculation
            idx = sorted_periods.index(key) if key in sorted_periods else -1
            prev_key = sorted_periods[idx - 1] if idx > 0 else None
            prev_bal = monthly_balances_full[prev_key] if prev_key else {}

            # ---- BALANCES (Full historical from ALL(Calendario)) ----
            cxc_end = sum(bal.get(c, 0) for c in classifications['cxc'])
            cxc_start = sum(prev_bal.get(c, 0) for c in classifications['cxc'])
            prom_cxc = (cxc_start + cxc_end) / 2

            inv_end = sum(bal.get(c, 0) for c in classifications['inventarios'])
            inv_start = sum(prev_bal.get(c, 0) for c in classifications['inventarios'])
            prom_inv = (inv_start + inv_end) / 2

            cxp_end = sum(bal.get(c, 0) for c in classifications['cxp_proveedores']) * -1
            cxp_start = sum(prev_bal.get(c, 0) for c in classifications['cxp_proveedores']) * -1
            prom_cxp = (cxp_start + cxp_end) / 2

            at_end = sum(bal.get(c, 0) for c in classifications['activos_totales'])
            at_start = sum(prev_bal.get(c, 0) for c in classifications['activos_totales'])
            prom_at = (at_start + at_end) / 2

            # ---- FLOWS (Operational, excluding Doc 998) ----
            ventas = sum(flows.get(c, 0) for c in classifications['ventas_41']) * -1
            costos_ventas = sum(flows.get(c, 0) for c in classifications['costos_ventas_6135'])
            compras = sum(flows.get(c, 0) for c in classifications['compras'])

            # ---- DYNAMIC CALENDAR (matches COUNTROWS('Calendario')) ----
            dias_periodo = calendar.monthrange(year, month)[1]

            # ---- INGRESOS OPERACIONALES (ALL flows, includes 998) ----
            # DAX: [Ingresos Operacionales] = [Neto] filtered by Grupo 41, NO 998 exclusion
            ingresos_op = sum(flows_all.get(c, 0) for c in classifications['ventas_41']) * -1

            # ---- ROTACIONES ----
            rot_cxc = safe_div(ventas, prom_cxc)
            rot_inv = safe_div(costos_ventas, prom_inv)
            rot_cxp = safe_div(compras, prom_cxp)
            rot_at = safe_div(ingresos_op, prom_at)  # Uses Ingresos Operacionales, NOT Ventas

            # ---- DÍAS ----
            dso = safe_div(dias_periodo, rot_cxc) if rot_cxc > 0 else 0
            dio = safe_div(dias_periodo, rot_inv) if rot_inv > 0 else 0
            dpo = safe_div(dias_periodo, rot_cxp) if rot_cxp > 0 else 0

            # ---- CICLO ----
            cce = dso + dio - dpo

            # ---- STORE RESULTS ----
            results['Días de Cartera (DSO)'].append((q, dso, month, year))
            results['Días de Inventario (DIO)'].append((q, dio, month, year))
            results['Días de Proveedores (DPO)'].append((q, dpo, month, year))
            results['Ciclo de Conversión de Efectivo'].append((q, cce, month, year))
            results['Rotación de Activos Totales'].append((q, rot_at, month, year))
            results['Rotación de Cartera'].append((q, rot_cxc, month, year))
            results['Rotación de Inventarios'].append((q, rot_inv, month, year))

    return results


# ============================================================
# STEP 6: Write CSVs
# ============================================================
INDICATORS = [
    'Ciclo de Conversión de Efectivo',
    'Días de Cartera (DSO)',
    'Días de Inventario (DIO)',
    'Días de Proveedores (DPO)',
    'Rotación de Activos Totales',
    'Rotación de Cartera',
    'Rotación de Inventarios',
]


def write_csvs(results):
    """Write indicator CSVs with the specified header format."""
    total = 0

    for indicator_name in INDICATORS:
        data = results.get(indicator_name, [])
        if not data:
            print(f"  WARNING: No data for {indicator_name}")
            continue

        filepath = os.path.join(OUTPUT_DIR, f"{indicator_name}.csv")

        with open(filepath, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(['Quarter', 'Days', 'Period', 'Año Texto'])
            for q, val, period, year in data:
                writer.writerow([q, val, period, year])

        total += 1
        print(f"  ✓ {indicator_name}.csv ({len(data)} rows)")

    print(f"\n  Written {total} CSV files to {OUTPUT_DIR}")
    return total


# ============================================================
# STEP 7: Compare with Power BI Ground Truth
# ============================================================
def compare_with_ground_truth(results):
    """Compare against original CSVs in sources/ACTIVIDAD/"""
    print("\n" + "=" * 60)
    print("VERIFICACIÓN CONTRA GROUND TRUTH (Power BI)")
    print("=" * 60)

    # Map acid test names to original file names
    name_map = {
        'Ciclo de Conversión de Efectivo': 'Ciclo de Conversión de Efectivo',
        'Días de Cartera (DSO)': 'Días de Cartera (DSO)',
        'Días de Inventario (DIO)': 'Días de Inventario (DIO)',
        'Días de Proveedores (DPO)': 'Días de Proveedores (DPO)',
        'Rotación de Activos Totales': 'Rotación de Activos Totales',
        'Rotación de Cartera': 'Rotación de Cartera',
        'Rotación de Inventarios': 'Rotación de Inventarios',
    }

    original_dir = os.path.join(SOURCES_DIR, 'ACTIVIDAD')
    total_comparisons = 0
    exact_matches = 0
    close_matches = 0

    for acid_name, orig_name in name_map.items():
        original_file = os.path.join(original_dir, f"{orig_name}.csv")
        if not os.path.exists(original_file):
            print(f"  [SKIP] {orig_name}.csv not found")
            continue

        # Read original values
        original_values = {}
        with open(original_file, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            headers = reader.fieldnames
            value_col = headers[1] if headers else 'Value'
            for row in reader:
                period = int(row['Period'])
                year = int(row['Año Texto'])
                try:
                    val = float(row[value_col].replace(',', '.').replace('"', ''))
                except (ValueError, KeyError):
                    continue
                original_values[(year, period)] = val

        # Compare
        ind_total = 0
        ind_exact = 0
        ind_close = 0
        mismatches = []

        generated = results.get(acid_name, [])
        for q, gen_val, period, year in generated:
            key = (year, period)
            if key not in original_values:
                continue

            orig_val = original_values[key]
            ind_total += 1
            total_comparisons += 1

            if orig_val == 0 and gen_val == 0:
                ind_exact += 1
                exact_matches += 1
            elif abs(gen_val - orig_val) < 0.01:
                ind_exact += 1
                exact_matches += 1
            elif orig_val != 0 and abs((gen_val - orig_val) / orig_val) < 0.05:
                ind_close += 1
                close_matches += 1
            else:
                mismatches.append((year, period, orig_val, gen_val))

        match_pct = (ind_exact + ind_close) / ind_total * 100 if ind_total > 0 else 0
        status = '✅' if match_pct >= 99.9 else '⚠️' if match_pct >= 80 else '❌'
        print(f"  {status} {acid_name}: {match_pct:.1f}% ({ind_exact}+{ind_close}/{ind_total})")

        if mismatches:
            for y, p, o, g in mismatches[:3]:
                print(f"      Mismatch [{y}-P{p}]: Orig={o:.4f} Gen={g:.4f}")

    good = exact_matches + close_matches
    print(f"\n  TOTAL: {good}/{total_comparisons} = {good/total_comparisons*100:.1f}% match" if total_comparisons > 0 else "")


# ============================================================
# MAIN
# ============================================================
def main():
    print("=" * 60)
    print("PRUEBA ÁCIDA — Módulo de Actividad")
    print("MAS CONSULTA SAS - NIT 901271750")
    print("=" * 60)

    print("\n[1/6] Loading Master Account...")
    accounts = load_account_classification()

    print("\n[2/6] Parsing movement files...")
    movements = parse_mov_files()

    print("\n[3/6] Building monthly balance snapshots...")
    m_full, m_flows_op, m_flows_all = build_monthly_balances(movements)

    print("\n[4/6] Classifying accounts for Activity indicators...")
    classifications = classify_accounts(accounts)

    print("\n[5/6] Calculating 7 Activity indicators...")
    results = calculate_activity_indicators(m_full, m_flows_op, classifications, m_flows_all)
    print(f"  Calculated {len(results)} indicators")

    print("\n[6/6] Writing CSVs...")
    write_csvs(results)

    # Verification
    compare_with_ground_truth(results)

    print("\n" + "=" * 60)
    print("PRUEBA ÁCIDA COMPLETADA")
    print("=" * 60)


if __name__ == '__main__':
    main()
