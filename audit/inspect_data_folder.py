"""Analisis profundo de Data/ vs DATOS ADICIONALES/ para entender diferencias."""
import sys, os
if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')

from openpyxl import load_workbook

DATA_DIR   = r'c:\Users\gutie\OneDrive\FERFILES\DATIA\PROPUESTAS\COLOMBIA\MASCONSULTA\Data'
DATOS_DIR  = r'c:\Users\gutie\OneDrive\FERFILES\DATIA\PROPUESTAS\COLOMBIA\MASCONSULTA\DATOS ADICIONALES'

print("="*70)
print("1. LAYOUT COMPARADO: Primeras 10 filas de datos en cada Mov 2024")
print("="*70)

def inspect_xlsx(fp, label, max_rows=12):
    print(f"\n--- {label} ---")
    wb = load_workbook(fp, read_only=True, data_only=True)
    ws = wb.active
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i >= max_rows:
            break
        # Solo mostrar filas no completamente vacias
        vals = [str(v)[:20] if v is not None else '' for v in row[:12]]
        if any(vals):
            print(f"  fila {i+1:3d}: {vals}")
    wb.close()

inspect_xlsx(os.path.join(DATA_DIR, '2024', 'Mov 2024.xlsx'), 'Data/2024/Mov 2024.xlsx')
inspect_xlsx(os.path.join(DATOS_DIR, 'Mov 2024.xlsx'), 'DATOS ADICIONALES/Mov 2024.xlsx')

print()
print("="*70)
print("2. PUC ADICIONAL.xlsx — primeras 20 filas significativas")
print("="*70)
puc_fp = os.path.join(DATA_DIR, 'PUC ADICIONAL.xlsx')
wb = load_workbook(puc_fp, read_only=True, data_only=True)
ws = wb.active
count = 0
for i, row in enumerate(ws.iter_rows(values_only=True)):
    vals = [str(v)[:25] if v is not None else '' for v in row[:9]]
    if any(vals):
        print(f"  fila {i+1:3d}: {vals}")
        count += 1
        if count >= 20:
            break
wb.close()
