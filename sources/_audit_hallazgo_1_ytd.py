from openpyxl import load_workbook
import csv
import os
from collections import defaultdict
from datetime import datetime

# Paths
BASE_DIR = r"c:\Users\gutie\OneDrive\FERFILES\DATIA\PROPUESTAS\COLOMBIA\+CONSULTA\LiquidityDashboard\sources"
MASTER_FILE = os.path.join(BASE_DIR, "Master Account.xlsx")
MOV_FILE = os.path.join(BASE_DIR, "Mov 2023.csv")

def audit_ytd():
    # 1. Load Classifications
    wb = load_workbook(MASTER_FILE, read_only=True, data_only=True)
    ws_accounts = wb['Accounts']
    ws_list = wb['List']
    
    group_termino = {}
    for row in ws_list.iter_rows(min_row=2, values_only=True):
        if row and len(row) >= 8:
            group_termino[str(row[5])] = str(row[7])

    ac_set = set()
    pc_set = set()
    
    for row in ws_accounts.iter_rows(min_row=2, values_only=True):
        if not row or not row[1]: continue
        codigo = str(row[1]).strip()
        clase = str(row[4])
        grupo = str(row[13])
        termino = group_termino.get(grupo, "")
        
        if codigo.startswith('1') and (termino == 'Corto Termino' or grupo in ('11', '12', '13', '14', '18', '19')):
            ac_set.add(codigo)
        if codigo.startswith('2') and (termino == 'Corto Termino' or grupo in ('22', '23', '24')):
            pc_set.add(codigo)
    
    # 2. YTD for 2023 (Jan-Feb)
    # AC = [Neto]
    # PC = [Neto] * -1
    
    ytd_neto = defaultdict(lambda: defaultdict(float))
    
    with open(MOV_FILE, 'r', encoding='utf-8', errors='replace') as f:
        reader = csv.reader(f)
        for row in reader:
            if not row or len(row) < 12 or not row[0].isdigit(): continue
            try:
                dt = datetime.strptime(row[1].strip(), '%d/%m/%Y')
                if dt.year == 2023:
                    month = dt.month
                    codigo = row[2].strip()
                    deb = float(row[10].replace(',', '')) if row[10] else 0.0
                    cre = float(row[11].replace(',', '')) if row[11] else 0.0
                    ytd_neto[month][codigo] += (deb - cre)
            except: continue
            
    for m in [1, 2]:
        print(f"\n--- PERIODO 2023-P{m} (YTD) ---")
        cur_ac = sum(sum(ytd_neto[i][c] for c in ac_set) for i in range(1, m+1))
        cur_pc_neto = sum(sum(ytd_neto[i][c] for c in pc_set) for i in range(1, m+1))
        cur_pc = cur_pc_neto * -1
        
        print(f"  Activo Corriente YTD: {cur_ac:15.2f}")
        print(f"  Pasivo Corriente YTD: {cur_pc:15.2f} (Neto: {cur_pc_neto:.2f})")
        if cur_pc != 0:
            print(f"  RATIO YTD: {cur_ac / cur_pc:.4f}")
        else:
            print("  RATIO YTD: N/A")

if __name__ == "__main__":
    audit_ytd()
