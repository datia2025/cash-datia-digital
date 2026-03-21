# Diagnostic: Show AC and PC breakdown by grupo for 2025-P1
import csv, os
from collections import defaultdict
from openpyxl import load_workbook
from datetime import datetime

SOURCES_DIR = r'c:\Users\gutie\OneDrive\FERFILES\DATIA\PROPUESTAS\COLOMBIA\+CONSULTA\LiquidityDashboard\sources'
MASTER_FILE = os.path.join(SOURCES_DIR, 'Master Account.xlsx')

# Load accounts
wb = load_workbook(MASTER_FILE, read_only=True, data_only=True)
ws_list = wb['List']
group_termino = {}
for row in ws_list.iter_rows(min_row=2, values_only=True):
    if row and len(row) >= 8:
        grupo = str(row[5]) if row[5] else None
        termino = str(row[7]) if row[7] else None
        if grupo and termino: group_termino[grupo] = termino

ws_accts = wb['Accounts']
accounts = {}
for row in ws_accts.iter_rows(min_row=2, values_only=True):
    if not row or not row[1]: continue
    codigo = str(row[1]).strip()
    grupo = str(row[13]) if row[13] else ''
    termino = str(row[18]) if len(row) > 18 and row[18] else group_termino.get(grupo, '')
    nombre_grupo = str(row[14]) if row[14] else ''
    accounts[codigo] = {'grupo': grupo, 'termino': termino, 'nombre_grupo': nombre_grupo}
wb.close()

# Parse all movements
movements = defaultdict(lambda: {'debito': 0.0, 'credito': 0.0})
for year in range(2021, 2026):
    filepath = os.path.join(SOURCES_DIR, f'Mov {year}.csv')
    if not os.path.exists(filepath): continue
    with open(filepath, 'r', encoding='utf-8') as f:
        for ln, line in enumerate(f):
            if ln < 7: continue
            try:
                reader = csv.reader([line])
                fields = next(reader)
                if len(fields) < 12: continue
                fecha_str = fields[1].strip()
                codigo = fields[2].strip()
                desc = fields[7].strip() if len(fields) > 7 else ''
                det = fields[8].strip() if len(fields) > 8 else ''
                if 'cierre anual' in desc.lower() or 'cierre anual' in det.lower(): continue
                fecha = datetime.strptime(fecha_str, '%d/%m/%Y')
                deb = float(fields[10].strip().replace(',','')) if fields[10].strip() else 0
                cre = float(fields[11].strip().replace(',','')) if fields[11].strip() else 0
                movements[(fecha.year, fecha.month, codigo)]['debito'] += deb
                movements[(fecha.year, fecha.month, codigo)]['credito'] += cre
            except: continue

# Build cumulative for target period
all_periods = sorted(set((y,m) for y,m,_ in movements.keys()))
running = defaultdict(float)
for year, month in all_periods:
    for (y,m,codigo), amounts in movements.items():
        if y != year or m != month: continue
        fd = codigo[0] if codigo else '0'
        if fd in ('1','5','6','7'):
            running[codigo] += amounts['debito'] - amounts['credito']
        elif fd in ('2','3','4'):
            running[codigo] += amounts['credito'] - amounts['debito']
    if (year, month) == (2025, 1):
        break

# Group the balances
ac_by_grupo = defaultdict(float)
pc_by_grupo = defaultdict(float)
for codigo, bal in running.items():
    fd = codigo[0]
    info = accounts.get(codigo, {'grupo': '??', 'termino': '', 'nombre_grupo': ''})
    g = info['grupo']
    ng = info['nombre_grupo'][:30]
    
    if fd == '1':
        if g in ('15','16','17'):
            pass  # no corriente
        else:
            ac_by_grupo[f"G{g} {ng}"] += bal
    elif fd == '2':
        pc_by_grupo[f"G{g} {ng}"] += bal

print('=== ACTIVO CORRIENTE by Grupo (2025-P1) ===')
total_ac = 0
for k in sorted(ac_by_grupo.keys()):
    v = ac_by_grupo[k]
    total_ac += v
    print(f"  {k}: {v:>20,.2f}")
print(f"  TOTAL AC: {total_ac:>20,.2f}")

print('\n=== PASIVO CORRIENTE by Grupo (2025-P1) ===')
total_pc = 0
for k in sorted(pc_by_grupo.keys()):
    v = pc_by_grupo[k]
    total_pc += v
    print(f"  {k}: {v:>20,.2f}")
print(f"  TOTAL PC: {total_pc:>20,.2f}")

print(f"\nRC = {total_ac/total_pc:.4f}")
print(f"CT = {total_ac-total_pc:,.2f}")
print(f"\nOriginal RC = -102.1554, Original CT = 23,707,902")
print(f"Implied original AC = 23,478,075, Implied original PC = -229,827")
print(f"\nDifference AC: {total_ac - 23478075:.2f} ({(total_ac - 23478075)/total_ac*100:.1f}%)")
print(f"Difference PC: {total_pc - (-229827):.2f}")
