from openpyxl import load_workbook
import csv
import os
from collections import defaultdict
from datetime import datetime

# Paths
BASE_DIR = r"c:\Users\gutie\OneDrive\FERFILES\DATIA\PROPUESTAS\COLOMBIA\+CONSULTA\LiquidityDashboard\sources"
MASTER_FILE = os.path.join(BASE_DIR, "Master Account.xlsx")
MOV_FILES = [os.path.join(BASE_DIR, f"Mov {y}.csv") for y in [2021, 2022, 2023]]

def audit_cumulative():
    # 1. Load Classifications
    wb = load_workbook(MASTER_FILE, read_only=True, data_only=True)
    ws_accounts = wb['Accounts']
    ws_list = wb['List']
    
    group_termino = {}
    for row in ws_list.iter_rows(min_row=2, values_only=True):
        if row and len(row) >= 8:
            group_termino[str(row[5])] = str(row[7])

    ac_set = set()
    pc_set = set()
    
    for row in ws_accounts.iter_rows(min_row=2, values_only=True):
        if not row or not row[1]: continue
        codigo = str(row[1]).strip()
        clase = str(row[4])
        grupo = str(row[13])
        termino = group_termino.get(grupo, "")
        
        if codigo.startswith('1') and (termino == 'Corto Termino' or grupo in ('11', '12', '13', '14', '18', '19')):
            ac_set.add(codigo)
        if codigo.startswith('2') and (termino == 'Corto Termino' or grupo in ('22', '23', '24')):
            pc_set.add(codigo)
    
    # 2. Cumulative Balances up to Feb 28 2023
    ac_bal = defaultdict(float)
    pc_bal = defaultdict(float)
    
    cutoff = datetime(2023, 2, 28)
    
    for mov_path in MOV_FILES:
        if not os.path.exists(mov_path): continue
        with open(mov_path, 'r', encoding='utf-8', errors='replace') as f:
            reader = csv.reader(f)
            for row in reader:
                if not row or len(row) < 12 or not row[0].isdigit(): continue
                try:
                    dt = datetime.strptime(row[1].strip(), '%d/%m/%Y')
                    if dt <= cutoff:
                        codigo = row[2].strip()
                        deb = float(row[10].replace(',', '')) if row[10] else 0.0
                        cre = float(row[11].replace(',', '')) if row[11] else 0.0
                        
                        # Neto = Debito - Credito
                        neto = deb - cre
                        
                        if codigo in ac_set:
                            ac_bal[codigo] += neto
                        if codigo in pc_set:
                            pc_bal[codigo] += neto
                except: continue
                
    print("--- ACTIVO CORRIENTE CUMULATIVE FEB 2023 ---")
    total_ac = 0
    for c, val in sorted(ac_bal.items(), key=lambda x: abs(x[1]), reverse=True):
        if abs(val) > 0.01:
            # DAX: Activo Corriente = [Neto]
            print(f"{c}: {val:15.2f}")
            total_ac += val
    print(f"TOTAL AC (Cumulative): {total_ac:15.2f}")

    print("\n--- PASIVO CORRIENTE CUMULATIVE FEB 2023 ---")
    total_pc = 0
    for c, val in sorted(pc_bal.items(), key=lambda x: abs(x[1]), reverse=True):
        if abs(val) > 0.01:
            # DAX: Pasivo Corriente = [Neto] * -1
            adj_val = val * -1
            print(f"{c}: {adj_val:15.2f} (Neto: {val:.2f})")
            total_pc += adj_val
    print(f"TOTAL PC (Cumulative): {total_pc:15.2f}")
    
    if total_pc != 0:
        print(f"\nRATIO: {total_ac / total_pc:.4f}")

if __name__ == "__main__":
    audit_cumulative()
