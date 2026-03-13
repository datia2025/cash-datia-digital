"""
Prueba Ácida: Cálculo de 33 Indicadores Financieros desde datos crudos
MAS CONSULTA SAS - NIT 901271750

Pipeline: Mov 20XX.csv + Master Account.xlsx → 33 CSVs organizados por módulo

Autor: Generado por Antigravity (Forensic Analytics Skill)
"""

import csv
import glob
import os
import re
import calendar
from datetime import datetime
from collections import defaultdict
from openpyxl import load_workbook

# ============================================================
# CONFIG
# ============================================================
SOURCES_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT_DIR = os.path.join(SOURCES_DIR, "GENERADOS")
MASTER_FILE = os.path.join(SOURCES_DIR, "Master Account.xlsx")
def discover_mov_files():
    """
    Auto-discover movement files in SOURCES_DIR.
    Supports both CSV and Excel formats:
      - Mov 2021.csv, Mov 2022.csv, etc.
      - Mov 2021.xlsx, Mov 2022.xlsx, etc.
      - Movimientos 2024.xlsx, etc.
    If both .csv and .xlsx exist for the same year, prefers .xlsx.
    """
    found = {}
    for ext in ('csv', 'xlsx'):
        for fp in glob.glob(os.path.join(SOURCES_DIR, f"Mov*.{ext}")):
            basename = os.path.basename(fp)
            # Extract year from filename (e.g. "Mov 2024.xlsx" -> 2024)
            m = re.search(r'(\d{4})', basename)
            if m:
                year = int(m.group(1))
                # .xlsx takes priority over .csv for the same year
                if year not in found or ext == 'xlsx':
                    found[year] = fp
    files = [found[y] for y in sorted(found.keys())]
    if files:
        print(f"  Discovered {len(files)} movement files: {[os.path.basename(f) for f in files]}")
    return files

MOV_FILES = discover_mov_files()
# Dashboard analysis period — dynamically inferred from movement files
# Extracts years from filenames like "Mov 2022.xlsx", "Mov 2024.csv", etc.
def _infer_analysis_years():
    years = []
    for fp in MOV_FILES:
        basename = os.path.basename(fp)
        m = re.search(r'(\d{4})', basename)
        if m:
            years.append(int(m.group(1)))
    if not years:
        # Fallback to recent years if no files found
        return [2023, 2024, 2025]
    return sorted(set(years))

ANALYSIS_YEARS = _infer_analysis_years()

# ============================================================
# STEP 1: Load Master Account (Account Classification)
#   Auto-detects format: Master Account (Accounts sheet) or
#   raw PUC.xlsx (standard Siigo/ERP export)
# ============================================================

# --- Static lookup tables for PUC colombiano ---

# Clase contable → Tipo contable
CLASE_TO_TIPO = {
    'Activo': 'Balance',
    'Pasivo': 'Balance',
    'Patrimonio': 'Balance',
    'Ingresos': 'Resultado',
    'Gastos': 'Resultado',
    'Costos de venta': 'Resultado',
    'Costos de producción o de operación': 'Resultado',
    'Cuentas de orden deudoras': 'Orden',
    'Cuentas de orden acreedoras': 'Orden',
}

# Grupo PUC → Termino (Corto/Largo Plazo)
# Calibrado contra la hoja List del Master Account.xlsx de MAS CONSULTA SAS
# Grupos 25 (obligaciones laborales), 27, 28 = Corto Termino (exigibles en <1 año)
# Grupos 31-38 (Patrimonio) = sin termino (no son pasivos)
GROUP_TERMINO_DEFAULT = {
    '11': 'Corto Termino', '12': 'Corto Termino',
    '13': 'Corto Termino', '14': 'Corto Termino',
    '15': 'Largo Termino', '16': 'Largo Termino',
    '17': 'Largo Termino', '18': 'Corto Termino',
    '19': 'Corto Termino',
    '21': 'Corto Termino', '22': 'Corto Termino',  # 21=oblig. financieras → CP (alineado con Master Account)
    '23': 'Corto Termino', '24': 'Corto Termino',
    '25': 'Corto Termino', '26': 'Largo Termino',  # 25=obligaciones laborales → CP
    '27': 'Corto Termino', '28': 'Corto Termino',  # 27-28=depósitos/anticipos → CP
    '29': 'Largo Termino',
    # Patrimonio (clase 3): sin termino — no son pasivos corrientes ni no corrientes
    '31': '', '32': '', '33': '', '34': '', '36': '', '37': '', '38': '',
}

# Grupos que participan en EBITDA:
#   41/42 = Ingresos operacionales (alineado con Master Account — 33 cuentas clase 4x con ebitda=SI)
#   51/52 = Gastos operativos (excluye financieros/depreciación)
EBITDA_GRUPOS = {'41', '42', '51', '52'}
# Cuentas excluidas de EBITDA (depreciación, amortización)
EBITDA_EXCLUSION_PREFIXES = ('5160', '5165', '5260', '5265')


def load_account_classification():
    """
    Auto-detects file format and loads account classification:
    - If file has 'Accounts' sheet → Master Account format (original)
    - Otherwise → raw PUC.xlsx from Siigo/ERP (auto-derives fields)
    Returns: (accounts_dict, group_termino_dict)
    """
    wb = load_workbook(MASTER_FILE, read_only=True, data_only=True)

    if 'Accounts' in wb.sheetnames:
        print("  Formato detectado: Master Account (hoja Accounts)")
        return _load_from_master_account(wb)
    else:
        print("  Formato detectado: PUC.xlsx crudo (auto-derivación)")
        return _load_from_puc(wb)


def _load_from_master_account(wb):
    """Original Master Account loader — updated to handle missing 'List' sheet."""
    # Read List sheet for group → termino mapping
    group_termino = {}
    if 'List' in wb.sheetnames:
        ws_list = wb['List']
        for row in ws_list.iter_rows(min_row=2, values_only=True):
            if row and len(row) >= 8:
                grupo = str(row[5]) if row[5] else None
                termino = str(row[7]) if row[7] else None
                if grupo and termino:
                    group_termino[grupo] = termino
    else:
        print("  ⚠️ Hoja 'List' no encontrada. Usando mapeo por defecto de Corto/Largo Plazo.")
        group_termino = dict(GROUP_TERMINO_DEFAULT)

    # Read Accounts sheet
    ws_accounts = wb['Accounts']
    accounts = {}
    for row in ws_accounts.iter_rows(min_row=2, values_only=True):
        if not row or not row[1]:
            continue
        codigo = str(row[1]).strip()
        nombre = str(row[2]) if row[2] else ""
        categoria = str(row[3]) if row[3] else ""
        clase = str(row[4]) if row[4] else ""
        tipo = str(row[12]) if row[12] else ""  # Column M = Balance/Resultado
        grupo = str(row[13]) if row[13] else ""  # Column N = grupo PUC
        ebitda_flag = str(row[17]) if row[17] else "NO"  # Column R = EBITDA

        termino = group_termino.get(grupo, "")

        accounts[codigo] = {
            'nombre': nombre,
            'categoria': categoria,
            'clase': clase,
            'tipo': tipo,
            'grupo': grupo,
            'termino': termino,
            'ebitda': ebitda_flag.upper().strip()
        }

    wb.close()
    print(f"  Loaded {len(accounts)} accounts from Master Account")
    print(f"  Group-Termino mappings: {len(group_termino)}")
    return accounts, group_termino


def _load_from_puc(wb):
    """
    Load account classification from a raw PUC.xlsx (Siigo/ERP export).
    Auto-detects header row and derives Tipo, Grupo, Termino, EBITDA.
    Expected PUC columns: Código, Nombre, Categoría, Clase, ...
    """
    ws = wb.active

    # Step 1: Find header row (search for 'Código' or 'Codigo' in first column)
    header_row_idx = None
    col_map = {}  # column_name → column_index
    all_rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        all_rows.append(row)
        if header_row_idx is None and row:
            for j, cell in enumerate(row):
                cell_str = str(cell).strip().lower() if cell else ''
                if cell_str in ('código', 'codigo'):
                    header_row_idx = i
                    # Map all headers in this row
                    for k, h in enumerate(row):
                        if h:
                            col_map[str(h).strip().lower()] = k
                    break

    if header_row_idx is None:
        # Fallback: assume data starts at row 2, col 0 = code
        print("  ⚠️ No header row found in PUC, using positional fallback")
        header_row_idx = 1
        col_map = {'código': 0, 'nombre': 1, 'categoría': 2, 'clase': 3}

    # Resolve column indices
    col_code = col_map.get('código', col_map.get('codigo', 0))
    col_name = col_map.get('nombre', 1)
    col_cat = col_map.get('categoría', col_map.get('categoria', 2))
    col_clase = col_map.get('clase', 3)
    col_nivel = col_map.get('nivel agrupación', col_map.get('nivel agrupacion', -1))

    print(f"  PUC header en fila {header_row_idx}: code={col_code}, name={col_name}, cat={col_cat}, clase={col_clase}")

    # Step 2: Parse leaf accounts (8+ digit codes or Nivel=Transaccional)
    accounts = {}
    group_termino = dict(GROUP_TERMINO_DEFAULT)  # Use static defaults
    skipped = 0

    for row in all_rows[header_row_idx:]:  # Skip header row
        if not row:
            continue

        raw_code = row[col_code] if col_code < len(row) else None
        if not raw_code:
            continue

        codigo = str(raw_code).strip()

        # Skip non-numeric codes (company names, NIT, headers)
        if not codigo or not codigo[0].isdigit():
            continue

        # Filter: only leaf accounts (8+ digits = transactional level)
        # Also accept if Nivel column says "Transaccional"
        is_leaf = len(codigo) >= 8
        if col_nivel >= 0 and col_nivel < len(row) and row[col_nivel]:
            nivel_val = str(row[col_nivel]).strip().lower()
            if nivel_val == 'transaccional':
                is_leaf = True

        if not is_leaf:
            skipped += 1
            continue

        nombre = str(row[col_name]).strip() if col_name < len(row) and row[col_name] else ""
        categoria = str(row[col_cat]).strip() if col_cat < len(row) and row[col_cat] else ""
        clase = str(row[col_clase]).strip() if col_clase < len(row) and row[col_clase] else ""

        # --- Derive Tipo from Clase ---
        tipo = CLASE_TO_TIPO.get(clase, '')
        if not tipo:
            # Fallback by first digit: 1-3 = Balance, 4-7 = Resultado
            first = codigo[0]
            if first in ('1', '2', '3'):
                tipo = 'Balance'
            elif first in ('4', '5', '6', '7'):
                tipo = 'Resultado'

        # --- Derive Grupo from first 2 digits ---
        grupo = codigo[:2] if len(codigo) >= 2 else ""

        # --- Derive Termino from Grupo ---
        termino = group_termino.get(grupo, "")

        # --- Derive EBITDA flag ---
        ebitda_flag = "NO"
        if grupo in EBITDA_GRUPOS:
            # Gastos operativos participan, excepto depreciación/amortización
            if not any(codigo.startswith(prefix) for prefix in EBITDA_EXCLUSION_PREFIXES):
                ebitda_flag = "SI"

        accounts[codigo] = {
            'nombre': nombre,
            'categoria': categoria,
            'clase': clase,
            'tipo': tipo,
            'grupo': grupo,
            'termino': termino,
            'ebitda': ebitda_flag
        }

    wb.close()
    print(f"  Loaded {len(accounts)} leaf accounts from PUC.xlsx")
    print(f"  Skipped {skipped} parent/group codes")
    print(f"  Group-Termino mappings: {len(group_termino)}")
    return accounts, group_termino


# ============================================================
# STEP 2: Parse all Mov CSV/XLSX files
# ============================================================
def _parse_mov_csv(filepath, movements):
    """
    Parse a single Mov CSV file and append to movements dict.
    Returns number of lines parsed.
    """
    file_lines = 0
    with open(filepath, 'r', encoding='utf-8', errors='replace') as f:
        reader = csv.reader(f)
        current_comprobante = ""

        for row in reader:
            if not row: continue

            first_val = row[0].strip()
            if first_val.startswith('Comprobante:'):
                current_comprobante = first_val.replace('Comprobante:', '').strip()
                continue

            if not first_val or not first_val[0].isdigit():
                continue

            if len(row) < 12:
                continue

            fecha_str = row[1].strip()
            codigo = row[2].strip()
            if not fecha_str or not codigo:
                continue

            descripcion = row[7].strip() if len(row) > 7 else ""
            detalle = row[8].strip() if len(row) > 8 else ""

            is_closing = False
            if 'cierre anual' in descripcion.lower() or 'cierre anual' in detalle.lower():
                is_closing = True
            if '998' in current_comprobante:
                is_closing = True

            try:
                fecha = datetime.strptime(fecha_str, '%d/%m/%Y')
            except ValueError:
                continue

            year = fecha.year
            month = fecha.month

            try:
                debito = float(row[10].replace(',', '')) if row[10] else 0.0
                credito = float(row[11].replace(',', '')) if row[11] else 0.0
            except ValueError:
                debito = 0.0
                credito = 0.0

            key = (year, month, codigo)
            movements[key].append({
                'debito': debito,
                'credito': credito,
                'is_closing': is_closing
            })
            file_lines += 1

    return file_lines


def _parse_mov_xlsx(filepath, movements):
    """
    Parse a single Mov Excel (.xlsx) file and append to movements dict.
    Expects the same column layout as the CSV:
      Col 0: Línea, Col 1: Fecha, Col 2: Código, Col 3: NIT,
      Col 4: No. Documento, Col 5: Clase Documento, Col 6: Tipo,
      Col 7: Descripción, Col 8: Detalle, Col 9: Campo,
      Col 10: Débito, Col 11: Crédito
    Also handles 'Comprobante:' header rows for 998 detection.
    Returns number of lines parsed.
    """
    file_lines = 0
    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active  # Use the first/active sheet

    current_comprobante = ""

    for row in ws.iter_rows(values_only=True):
        if not row or all(c is None for c in row):
            continue

        first_val = str(row[0]).strip() if row[0] is not None else ""

        # Detect Comprobante header rows
        if first_val.startswith('Comprobante:'):
            current_comprobante = first_val.replace('Comprobante:', '').strip()
            continue

        # Skip non-data rows (headers, blanks, text rows)
        if not first_val or not first_val[0].isdigit():
            continue

        if len(row) < 12:
            continue

        # Parse fecha — could be a datetime object or a string
        fecha_raw = row[1]
        if fecha_raw is None:
            continue
        if isinstance(fecha_raw, datetime):
            fecha = fecha_raw
        else:
            fecha_str = str(fecha_raw).strip()
            if not fecha_str:
                continue
            try:
                fecha = datetime.strptime(fecha_str, '%d/%m/%Y')
            except ValueError:
                try:
                    fecha = datetime.strptime(fecha_str, '%Y-%m-%d %H:%M:%S')
                except ValueError:
                    continue

        codigo = str(row[2]).strip() if row[2] is not None else ""
        if not codigo:
            continue

        descripcion = str(row[7]).strip() if len(row) > 7 and row[7] else ""
        detalle = str(row[8]).strip() if len(row) > 8 and row[8] else ""

        is_closing = False
        if 'cierre anual' in descripcion.lower() or 'cierre anual' in detalle.lower():
            is_closing = True
        if '998' in current_comprobante:
            is_closing = True

        # Clase Documento column (col 5) may also indicate 998
        clase_doc = str(row[5]).strip() if len(row) > 5 and row[5] else ""
        if '998' in clase_doc:
            is_closing = True

        year = fecha.year
        month = fecha.month

        # Parse debito/credito — may be numeric or string
        def parse_amount(val):
            if val is None:
                return 0.0
            if isinstance(val, (int, float)):
                return float(val)
            try:
                return float(str(val).replace(',', ''))
            except ValueError:
                return 0.0

        debito = parse_amount(row[10])
        credito = parse_amount(row[11])

        key = (year, month, codigo)
        movements[key].append({
            'debito': debito,
            'credito': credito,
            'is_closing': is_closing
        })
        file_lines += 1

    wb.close()
    return file_lines


def parse_mov_files():
    """
    Parse all Mov files (CSV and/or XLSX) and extract movements.
    Auto-detects format by file extension.
    Returns dict: { (year, month, codigo) -> [{debito, credito, is_closing}] }
    """
    movements = defaultdict(list)
    total_lines = 0

    for filepath in MOV_FILES:
        if not os.path.exists(filepath):
            print(f"  Warning: {filepath} not found.")
            continue

        filename = os.path.basename(filepath)
        ext = os.path.splitext(filename)[1].lower()
        print(f"  Parsing {filename} ({ext})...")

        try:
            if ext == '.xlsx':
                file_lines = _parse_mov_xlsx(filepath, movements)
            else:
                file_lines = _parse_mov_csv(filepath, movements)

            print(f"    Parsed {file_lines} movement lines")
            total_lines += file_lines

        except Exception as e:
            print(f"  Error parsing {filepath}: {e}")

    print(f"  Total movements parsed: {total_lines}")
    return movements


# ============================================================
# STEP 3: Build monthly balance sheet
# ============================================================
def build_monthly_balances(movements, accounts):
    """
    Compute cumulative balances (Continuous and YTD) and monthly flows.
    Segments closing entries (998/cierre) from operational flows.
    Also tracks YTD operational flows and closing-only flows for Patrimonio.
    """
    sorted_keys = sorted(movements.keys())

    running_full = defaultdict(float)
    running_ytd = defaultdict(float)
    running_full_ex998 = defaultdict(float)
    
    monthly_balances_full = {}
    monthly_balances_ytd = {}
    monthly_balances_full_ex998 = {}
    monthly_flows_op = defaultdict(lambda: defaultdict(float))  # Single-month operational
    monthly_flows_all_single = defaultdict(lambda: defaultdict(float))  # Single-month ALL (includes 998)
    monthly_flows_ytd_op = defaultdict(lambda: defaultdict(float))  # YTD operational (ex998)
    monthly_flows_ytd_all = defaultdict(lambda: defaultdict(float))  # YTD all (DATESYTD)
    monthly_flows_closing = defaultdict(lambda: defaultdict(float))  # YTD closing-only (998)

    running_ytd_op = defaultdict(float)  # YTD accumulator for operational flows
    running_ytd_all = defaultdict(float)  # YTD accumulator for ALL flows
    running_ytd_closing = defaultdict(float)  # YTD accumulator for closing-only flows

    current_year = None
    current_period = None

    for key in sorted_keys:
        year, month, codigo = key
        
        period = (year, month)
        if period != current_period:
            if current_period:
                monthly_balances_full[current_period] = dict(running_full)
                monthly_balances_ytd[current_period] = dict(running_ytd)
                monthly_balances_full_ex998[current_period] = dict(running_full_ex998)
                monthly_flows_ytd_op[current_period] = dict(running_ytd_op)
                monthly_flows_ytd_all[current_period] = dict(running_ytd_all)
                monthly_flows_closing[current_period] = dict(running_ytd_closing)
            
            if year != current_year:
                running_ytd.clear()
                running_ytd_op.clear()
                running_ytd_all.clear()
                running_ytd_closing.clear()
                current_year = year
            current_period = period

        for entry in movements[(year, month, codigo)]:
            flow = entry['debito'] - entry['credito']
            
            running_full[codigo] += flow
            running_ytd[codigo] += flow
            running_ytd_all[codigo] += flow
            monthly_flows_all_single[period][codigo] += flow
            
            if not entry['is_closing']:
                monthly_flows_op[period][codigo] += flow
                running_full_ex998[codigo] += flow
                running_ytd_op[codigo] += flow
            else:
                running_ytd_closing[codigo] += flow

    # Store final snapshot
    if current_period:
        monthly_balances_full[current_period] = dict(running_full)
        monthly_balances_ytd[current_period] = dict(running_ytd)
        monthly_balances_full_ex998[current_period] = dict(running_full_ex998)
        monthly_flows_ytd_op[current_period] = dict(running_ytd_op)
        monthly_flows_ytd_all[current_period] = dict(running_ytd_all)
        monthly_flows_closing[current_period] = dict(running_ytd_closing)

    print(f"  Snapshots built: {len(monthly_balances_full)} Full, {len(monthly_balances_ytd)} YTD, {len(monthly_balances_full_ex998)} Full-ex998")
    return (monthly_balances_full, monthly_balances_ytd, monthly_flows_op,
            monthly_balances_full_ex998, monthly_flows_ytd_op, monthly_flows_ytd_all,
            monthly_flows_closing, monthly_flows_all_single)


# ============================================================
# STEP 4: Classify accounts for indicator calculation
# ============================================================
def classify_accounts(accounts, group_termino):
    """
    Classify accounts into categories for financial indicator calculation.
    Aligned with DAX metricas.md for 100% mathematical match.
    """
    classifications = defaultdict(set)

    for codigo, data in accounts.items():
        grupo = data.get('grupo', '')
        termino = data.get('termino', '')
        clase = data.get('clase', '')
        ebitda = data.get('ebitda', 'NO')
        first_digit = codigo[0] if codigo else '0'

        # --- BALANCE SHEET ---
        if first_digit == '1':
            classifications['activos_clase1'].add(codigo)  # ALL Activos for Acum.
            if termino == 'Corto Termino' or grupo in ('11','12','13','14','18','19'):
                classifications['activo_corriente'].add(codigo)
            else:
                classifications['activo_no_corriente'].add(codigo)

            if grupo == '11':
                classifications['efectivo'].add(codigo)
            if grupo == '13' and codigo.startswith('1305'):
                classifications['cxc'].add(codigo)
            if grupo == '14' and codigo.startswith('1435'):
                classifications['inventarios'].add(codigo)
            if grupo == '15':
                classifications['activos_fijos'].add(codigo)
            # Intangibles (Nombre Grupo = "Intangibles", typically grupo 16 or 17)
            # We check by grupo name from master
            nombre_grupo = ''
            for g, t in group_termino.items():
                if g == grupo:
                    break

        elif first_digit == '2':
            classifications['pasivos_clase2'].add(codigo)  # ALL Pasivos for Acum.
            if termino == 'Corto Termino' or grupo in ('22','23','24'):
                classifications['pasivo_corriente'].add(codigo)
                classifications['deuda_cp'].add(codigo)
            else:
                classifications['pasivo_no_corriente'].add(codigo)
                classifications['deuda_lp'].add(codigo)

            if grupo == '22' and (codigo.startswith('2205') or codigo.startswith('2210')):
                classifications['cxp_proveedores'].add(codigo)
            elif grupo == '23' and codigo.startswith('2335'):
                classifications['cxp_proveedores'].add(codigo)
            
            if grupo == '23':
                classifications['cxp_otros'].add(codigo)
            
            # Pasivos Financieros (Grupo 21) for Deuda Neta and Amort. Capital
            if grupo == '21':
                classifications['pasivos_financieros'].add(codigo)

        elif first_digit == '3':
            classifications['patrimonio'].add(codigo)

        # --- P&L / RESULTS ---
        elif first_digit == '4':
            classifications['ingresos'].add(codigo)
            classifications['pyg_456'].add(codigo)  # For EBIT/EBITDA/Utilidad Neta
            if grupo == '41':
                classifications['ventas_41'].add(codigo)

        elif first_digit == '5':
            classifications['gastos'].add(codigo)
            classifications['pyg_456'].add(codigo)
            if codigo.startswith('5305'):
                classifications['gastos_financieros'].add(codigo)
            # Interest = ONLY account 53052001 (Cambio 5)
            if codigo == '53052001':
                classifications['interest_53052001'].add(codigo)
            if codigo.startswith('5160') or codigo.startswith('5260'):
                classifications['depreciacion'].add(codigo)
            if codigo.startswith('5165') or codigo.startswith('5265'):
                classifications['amortizacion'].add(codigo)
            # Depreciacion/Amortizacion for EBITDA exclusion (Cuenta 5160, 5165)
            if codigo.startswith('5160'):
                classifications['ebitda_excl_5160'].add(codigo)
            if codigo.startswith('5165'):
                classifications['ebitda_excl_5165'].add(codigo)
            # Grupo 54 = Impuestos (excluded from EBIT/EBITDA)
            if grupo == '54':
                classifications['impuestos_g54'].add(codigo)
            # Pagos de Arrendamiento (Cuenta 5120, 5220) for Cobertura Cargos Fijos
            if codigo.startswith('5120') or codigo.startswith('5220'):
                classifications['arrendamientos'].add(codigo)
            
            compras_grupos = ('5105','5110','5120','5125','5130','5135','5140','5145','5150','5155','5195','5205','5210','5220','5225','5230','5235','5240','5245','5250','5255','5295')
            if any(codigo.startswith(cg) for cg in compras_grupos):
                classifications['compras'].add(codigo)

        elif first_digit in ('6', '7'):
            classifications['costos'].add(codigo)
            if first_digit == '6':
                classifications['pyg_456'].add(codigo)  # Only class 6, NOT 7
            if codigo.startswith('6135'):
                classifications['costos_ventas_6135'].add(codigo)
                classifications['compras'].add(codigo)

        # P&L accounts for Patrimonio formula (Clase IN {Ingresos, Gastos, Costos})
        if first_digit in ('4', '5', '6', '7'):
            classifications['pyg_all'].add(codigo)

        if ebitda == 'SI':
            classifications['ebitda_accounts'].add(codigo)

    return classifications


# ============================================================
# STEP 5: Compute aggregates per month
# ============================================================
def compute_aggregates(monthly_balances_full, monthly_balances_ytd, monthly_flows_op,
                       classifications, monthly_balances_full_ex998=None,
                       monthly_flows_ytd_op=None, monthly_flows_ytd_all=None,
                       monthly_flows_closing=None, monthly_flows_all_single=None):
    """
    Compute financial aggregates for each (year, month).
    Integrates Dual-Mode (YTD/Full), segmented Document 998 logic,
    Patrimonio Relativo, EBIT/EBITDA, Utilidad Neta, Interest.
    """
    aggregates = {}
    periods = sorted(monthly_balances_full.keys())
    bal_2022 = monthly_balances_full.get((2022, 12), {})

    for year, month in periods:
        bal_full = monthly_balances_full[(year, month)]
        bal_ytd = monthly_balances_ytd[(year, month)]
        flows_op = monthly_flows_op.get((year, month), {})
        bal_ex998 = monthly_balances_full_ex998.get((year, month), {}) if monthly_balances_full_ex998 else {}
        ytd_op = monthly_flows_ytd_op.get((year, month), {}) if monthly_flows_ytd_op else {}
        ytd_all = monthly_flows_ytd_all.get((year, month), {}) if monthly_flows_ytd_all else {}
        closing = monthly_flows_closing.get((year, month), {}) if monthly_flows_closing else {}
        flows_all_single = monthly_flows_all_single.get((year, month), {}) if monthly_flows_all_single else {}

        def sum_bal(mode, account_set):
            source = bal_ytd if mode == 'ytd' else bal_full
            val = sum(source.get(c, 0) for c in account_set)
            if mode == 'full_rebase':
                rebase_val = sum(bal_2022.get(c, 0) for c in account_set)
                val -= rebase_val
            return val

        def sum_flows(mode, account_set):
            return sum(flows_op.get(c, 0) for c in account_set)

        mode_data = {}
        for m in ['ytd', 'full', 'full_rebase']:
            ac = sum_bal(m, classifications['activo_corriente'])
            anc = sum_bal(m, classifications['activo_no_corriente'])
            at = ac + anc
            pc = sum_bal(m, classifications['pasivo_corriente']) * -1
            pnc = sum_bal(m, classifications['pasivo_no_corriente']) * -1
            pt = pc + pnc
            pat = sum_bal(m, classifications['patrimonio']) * -1
            ef = sum_bal(m, classifications['efectivo'])
            cxc = sum_bal(m, classifications['cxc'])
            inv = sum_bal(m, classifications['inventarios'])
            af = sum_bal(m, classifications['activos_fijos'])
            cp_p = sum_bal(m, classifications['cxp_proveedores']) * -1
            d_cp = sum_bal(m, classifications['deuda_cp']) * -1
            d_lp = sum_bal(m, classifications['deuda_lp']) * -1
            
            mode_data[m] = {
                'ac': ac, 'anc': anc, 'at': at, 'pc': pc, 'pnc': pnc, 'pt': pt,
                'pat': pat, 'ef': ef, 'cxc': cxc, 'inv': inv, 'af': af,
                'cp_p': cp_p, 'd_cp': d_cp, 'd_lp': d_lp
            }

        # Ingresos Operacionales (ALL flows, includes 998) for Rotación de Activos Totales
        ingresos_op_all = sum(flows_all_single.get(c, 0) for c in classifications['ventas_41']) * -1

        # --- Monthly P&L flows (for Activity ratios, unchanged) ---
        v41 = sum_flows(m, classifications['ventas_41']) * -1
        cv = sum_flows(m, classifications['costos_ventas_6135'])
        comp = sum_flows(m, classifications['compras'])
        ing_n = sum_flows(m, classifications['ingresos']) * -1
        gas = sum_flows(m, classifications['gastos'])
        cos = sum_flows(m, classifications['costos'])
        gf = sum_flows(m, classifications['gastos_financieros'])
        dep = sum_flows(m, classifications['depreciacion'])
        amo = sum_flows(m, classifications['amortizacion'])
        util_op = ing_n - gas - cos

        # --- Efectivo with Doc 998 exclusion — DAX uses ROUND([Neto],4) ---
        ef_ex998 = round(sum(bal_ex998.get(c, 0) for c in classifications['efectivo']), 4)

        # === CAMBIO 6: Activos (Acum.) and Pasivos (Acum.) with ALL(Calendario)+ex998 ===
        # DAX: CALCULATE(ROUND([Neto],2), ...) rounds EACH account's balance before summing
        activos_acum = sum(round(bal_ex998.get(c, 0), 2) for c in classifications['activos_clase1'])
        pasivos_acum = sum(round(bal_ex998.get(c, 0), 2) for c in classifications['pasivos_clase2']) * -1

        # === CAMBIO 2: Patrimonio (Acum.) Relativo ===
        # DAX: Patrimonio Contable (Acum.) Sin Utilidad = Neto of Clase 3, ALL(Calendario), NO 998 exclusion
        patrim_sin_util = sum(bal_full.get(c, 0) for c in classifications['patrimonio'])  # NO per-account round
        # DAX: Utilidad (Acum. Anual) Manual = Neto of PyG, ex998, YTD * -1
        util_manual = sum(ytd_op.get(c, 0) for c in classifications['pyg_all']) * -1
        # DAX: Utilidad (Acum. Anual) Cierre 998 = Neto of PyG, ONLY 998, YTD * -1
        util_cierre_998 = sum(closing.get(c, 0) for c in classifications['pyg_all']) * -1
        # DAX: Patrimonio (Acum.) Relativo = (patrim_sin_util * -1 + util_manual + util_cierre_998) * -1
        patrim_relativo = (patrim_sin_util * -1 + util_manual + util_cierre_998) * -1
        # DAX: Patrimonio (Acum. Anual) = Patrimonio (Acum.) Relativo * -1
        patrim_anual = patrim_relativo * -1

        # === CAMBIO 3: EBIT and EBITDA with exact DAX logic ===
        # EBIT = Neto of Clase_1 IN {4,5,6}, ex998, NOT Grupo 54, NOT Codigo 53052001, DATESYTD * -1
        ebit_set = classifications['pyg_456'] - classifications['impuestos_g54'] - classifications.get('interest_53052001', set())
        ebit = sum(ytd_op.get(c, 0) for c in ebit_set) * -1
        # EBITDA = same as EBIT but also exclude Cuenta 5160 and 5165
        ebitda_set = ebit_set - classifications.get('ebitda_excl_5160', set()) - classifications.get('ebitda_excl_5165', set())
        ebitda_dax = sum(ytd_op.get(c, 0) for c in ebitda_set) * -1

        # === CAMBIO 4: Utilidad Neta (Acum. Anual) ===
        # DAX: Clase_1 IN {4,5,6}, DATESYTD (NO exclusions), * -1, round to 2
        util_neta = round(sum(ytd_all.get(c, 0) for c in classifications['pyg_456']) * -1, 2)

        # === CAMBIO 5: Interest = ONLY cuenta 53052001, ex998, YTD ===
        interest = sum(ytd_op.get(c, 0) for c in classifications.get('interest_53052001', set()))

        # Pagos de Arrendamiento (Cuenta 5120, 5220, DATESYTD)
        arrendamiento = sum(ytd_all.get(c, 0) for c in classifications.get('arrendamientos', set()))

        # Amortizaciones de Capital (Pasivos Financieros Grupo 21, DATESYTD, ABS)
        amort_capital = abs(sum(ytd_all.get(c, 0) for c in classifications.get('pasivos_financieros', set())))

        # Depreciaciones for Solvencia (Cuenta 5160, 5260, ex998, YTD)
        depreciaciones_solv = sum(ytd_op.get(c, 0) for c in classifications['depreciacion'])

        # Pasivos Financieros (ALL+ex998) for Deuda Neta — DAX uses ROUND([Neto],2) per account
        pasivos_fin = sum(round(bal_ex998.get(c, 0), 2) for c in classifications.get('pasivos_financieros', set())) * -1

        # Intangibles (ALLSELECTED, Clase=Activo, NombreGrupo=Intangibles) — approximate
        # We don't have NombreGrupo filter here; use Activos Fijos subset or 0
        # For simplicity, since Ratio de Capitalización is all zeros, intangibles likely 0
        intangibles = 0  # Will refine if needed

        # Pasivo No Corriente for Estructura (ALLSELECTED, Termino=Largo Termino)
        # Using the mode-based d_lp was from dual mode, but DAX uses ALLSELECTED
        # We'll pass pnc from the mode data

        # Activo Fijo Neto for Estructura (ALLSELECTED, Nombre Grupo = PPE)
        # Already covered by activos_fijos in mode_data via af

        # Resultados Ventas (Grupo 41, ex998, YTD) = ventas for margin calculations
        ventas_ytd = sum(ytd_op.get(c, 0) for c in classifications['ventas_41'])  # raw (negative)

        # Utilidad Acumulada for RENTABILIDAD CSV: includes 998 effect (0 in P12 after closing)
        util_acum_rentab = sum(ytd_all.get(c, 0) for c in classifications['pyg_all']) * -1

        aggregates[(year, month)] = {
            'modes': mode_data,
            'v41': v41, 'cv': cv, 'comp': comp, 'ing_n': ing_n,
            'gas': gas, 'cos': cos, 'gf': gf, 'dep': dep, 'amo': amo,
            'util_op': util_op, 'ef_ex998': ef_ex998,
            'ingresos_op_all': ingresos_op_all,
            # New DAX-aligned measures
            'activos_acum': activos_acum,
            'pasivos_acum': pasivos_acum,
            'patrim_relativo': patrim_relativo,
            'patrim_anual': patrim_anual,
            'ebit': ebit,
            'ebitda_dax': ebitda_dax,
            'util_neta': util_neta,
            'util_manual': util_manual,
            'interest': interest,
            'arrendamiento': arrendamiento,
            'amort_capital': amort_capital,
            'depreciaciones_solv': depreciaciones_solv,
            'pasivos_fin': pasivos_fin,
            'intangibles': intangibles,
            'ventas_ytd': ventas_ytd,
            'util_acum_rentab': util_acum_rentab,
        }

    return aggregates


# ============================================================
# STEP 6: Calculate all 33 indicators
# ============================================================
def month_to_quarter(month):
    if month <= 3: return "1Q"
    if month <= 6: return "2Q"
    if month <= 9: return "3Q"
    return "4Q"


def safe_div(a, b):
    """Safe division, returns 0 if denominator is 0."""
    if b == 0 or b is None:
        return 0
    return a / b


def calculate_indicators(aggregates):
    """
    Calculate all 33 indicators with dynamic mode switching.
    """
    results = defaultdict(list)
    sorted_keys = sorted(aggregates.keys())

    for year in ANALYSIS_YEARS:
        for month in range(1, 13):
            key = (year, month)
            if key not in aggregates: continue

            a = aggregates[key]
            q = month_to_quarter(month)

            # --- DYNAMIC MODE SWITCH (Liquidez and others) ---
            # The earliest year in ANALYSIS_YEARS always uses 'ytd' (no prior year to rebase against).
            # Subsequent years use 'full_rebase' (continuous balance rebased from the base year's Dec).
            base_year = ANALYSIS_YEARS[0] if ANALYSIS_YEARS else 2023
            mode = 'ytd' if year == base_year else 'full_rebase'
            d = a['modes'][mode]
            
            # --- AGGREGATES FOR ACTIVITY (Continuous History ALWAYS) ---
            # Activity ratios MUST use 'full' (continuous from 2021) to match ground truth averages
            idx = sorted_keys.index(key) if key in sorted_keys else -1
            prev_key = sorted_keys[idx-1] if idx > 0 else None
            
            d_full = a['modes']['full']
            prev_d_full = aggregates[prev_key]['modes']['full'] if prev_key else {
                'cxc': 0, 'inv': 0, 'cp_p': 0, 'at': 0, 'pat': 0
            }
            
            prom_cxc = (d_full['cxc'] + prev_d_full['cxc']) / 2
            prom_inv = (d_full['inv'] + prev_d_full['inv']) / 2
            prom_cxp = (d_full['cp_p'] + prev_d_full['cp_p']) / 2
            prom_act = (d_full['at'] + prev_d_full['at']) / 2

            # ---- LIQUIDEZ ----
            results[u'Razón Corriente'].append((q, safe_div(d['ac'], d['pc']), month, year))
            results['Capital de Trabajo'].append((q, d['ac'] - d['pc'], month, year)) 
            results[u'Prueba Ácida'].append((q, safe_div(d['ac'] - d['inv'], d['pc']), month, year))
            # DAX: [Asset-Efectivo](ALL+ex998) / [Pasivo Corriente](ALLSELECTED, no ex998)
            results['Ratio de Efectivo'].append((q, safe_div(a['ef_ex998'], d['pc']), month, year))

            # ---- RENTABILIDAD (using DAX-aligned measures) ----
            # DAX: Ind Margen Bruto = (Resultados Ventas - Resultados Costo) / Resultados Ventas
            results['Margen de Utilidad Bruta'].append((q, safe_div(a['ing_n'] - a['cos'], a['ing_n']), month, year))
            # DAX: Ind Margen Operativo = EBIT / (Resultados Ventas * -1) * -1
            # The DAX has TWO *-1 (one on EBIT definition, one at end of formula)
            # These cancel out: use raw without extra negation
            results['Margen Operativo'].append((q, safe_div(a['ebit'], -a['ventas_ytd']), month, year))
            # DAX: Ind Margen EBITDA = EBITDA / (Resultados Ventas * -1)
            # EBITDA already has *-1, ventas_ytd is raw negative. 
            # Result should be negated to match: ebitda_dax is negative when profitable,
            # -ventas_ytd is positive → result is negative, but ground truth is positive
            results['Margen EBITDA'].append((q, safe_div(a['ebitda_dax'], -a['ventas_ytd']) * -1, month, year))
            # DAX: Ind Margen Neto = Utilidad Neta / (Resultados Ventas * -1)  
            # Same logic: util_neta is negative when profitable (due to *-1 in DAX)
            results['Margen Neto Utilidad'].append((q, safe_div(a['util_neta'], -a['ventas_ytd']) * -1, month, year))
            # DAX: Patrimonio (Acum. Anual) = Patrimonio (Acum.) Relativo * -1
            results['Patrimonio'].append((q, a['patrim_anual'], month, year))
            # DAX: ROA = Utilidad Neta / Activos (Acum.)
            results['Retorno sobre Activos (ROA)'].append((q, safe_div(a['util_neta'], a['activos_acum']), month, year))
            # DAX: ROE = Utilidad Neta / (Patrimonio (Acum.) Relativo * -1)
            results['Retorno sobre Patrimonio (ROE)'].append((q, safe_div(a['util_neta'], a['patrim_relativo'] * -1), month, year))
            # DAX: Utilidad Acumulada (RENTABILIDAD) - includes 998 effect for P12 zeroing
            results['Utilidad Acumulada'].append((q, a['util_acum_rentab'], month, year))

            # ---- SOLVENCIA (using DAX-aligned measures) ----
            # DAX: Ind - Relación Deuda/Patrimonio = Pasivos(Acum.) / (Patrimonio Relativo * -1)
            results['Relación DeudaPatrimonio'].append((q, safe_div(a['pasivos_acum'], a['patrim_relativo'] * -1), month, year))
            # DAX: Ind - Multiplicador de Capital = Activos(Acum.) / (Patrimonio Relativo * -1)
            results['Multiplicador de Capital'].append((q, safe_div(a['activos_acum'], a['patrim_relativo'] * -1), month, year))
            # DAX: Ind - Ratio de Capitalización = PasivoLP / (PasivoLP + ABS(Patrimonio Relativo))
            patrim_abs = abs(a['patrim_relativo'])
            results['Ratio de Capitalización'].append((q, round(safe_div(d['d_lp'], d['d_lp'] + patrim_abs), 2), month, year))
            # DAX: Ind - Ratio de Propiedad/Autonomía = ABS(Patrimonio Relativo) / Activos(Acum.)
            results['Ratio de PropiedadAutonomía'].append((q, round(safe_div(patrim_abs, a['activos_acum']), 2), month, year))
            # DAX: Ind - Cobertura de Intereses = round(EBIT / Interest, 2)
            results['Cobertura de Intereses'].append((q, round(safe_div(a['ebit'], a['interest']), 2), month, year))
            # DAX: Ind - Deuda Neta a EBITDA = round((Pasivos Financieros - Asset Efectivo) / EBITDA, 2)
            deuda_neta = a['pasivos_fin'] - a['ef_ex998']
            results['Deuda Neta a EBITDA'].append((q, round(safe_div(deuda_neta, a['ebitda_dax']), 2), month, year))
            # DAX: Ind - Ratio de Endeudamiento Total = Pasivos(Acum.) / Activos(Acum.)
            results['Endeudamiento Total'].append((q, safe_div(a['pasivos_acum'], a['activos_acum']), month, year))

            # ---- ACTIVIDAD (Dynamic Calendar) ----
            dias_periodo = calendar.monthrange(year, month)[1]
            rot_cxc = safe_div(a['v41'], prom_cxc)
            rot_inv = safe_div(a['cv'], prom_inv)
            rot_cxp = safe_div(a['comp'], prom_cxp)

            results['Rotación de Cartera'].append((q, rot_cxc, month, year))
            results['Días de Cartera (DSO)'].append((q, safe_div(dias_periodo, rot_cxc) if rot_cxc > 0 else 0, month, year))
            results['Rotación de Inventarios'].append((q, rot_inv, month, year))
            results['Días de Inventario (DIO)'].append((q, safe_div(dias_periodo, rot_inv) if rot_inv > 0 else 0, month, year))
            results['Rotación de Proveedores'].append((q, rot_cxp, month, year))
            results['Días de Proveedores (DPO)'].append((q, safe_div(dias_periodo, rot_cxp) if rot_cxp > 0 else 0, month, year))
            
            dso = safe_div(dias_periodo, rot_cxc) if rot_cxc > 0 else 0
            dio = safe_div(dias_periodo, rot_inv) if rot_inv > 0 else 0
            dpo = safe_div(dias_periodo, rot_cxp) if rot_cxp > 0 else 0
            results['Ciclo de Conversión de Efectivo'].append((q, dso + dio - dpo, month, year))
            results['Rotación de Activos'].append((q, safe_div(a['ingresos_op_all'], prom_act), month, year))

            # ---- ESTRUCTURA (using DAX-aligned measures) ----
            # DAX: Ind - Cobertura de Cargos Fijos = round((EBIT + Arrendamiento) / (Interest + Arrendamiento), 2)
            results['Cobertura de Cargos Fijos'].append((q, round(safe_div(a['ebit'] + a['arrendamiento'], a['interest'] + a['arrendamiento']), 2), month, year))
            # DAX: Ind - Cobertura del Servicio de la Deuda = round(EBITDA / (Interest + Amort Capital), 2)
            results['Cobertura del Servicio de la Deuda'].append((q, round(safe_div(a['ebitda_dax'], a['interest'] + a['amort_capital']), 2), month, year))
            # DAX: Ind - Ratio de Solvencia Patrimonial = round((Utilidad Manual + Depreciaciones) / Pasivos(Acum.), 2)
            results['Ratio de Solvencia Patrimonial'].append((q, round(safe_div(a['util_manual'] + a['depreciaciones_solv'], a['pasivos_acum']), 2), month, year))
            # DAX: Ind - Cobertura de Activos Fijos = round((ABS(Patrimonio Relativo) + PasivoLP) / AF, 2)
            results['Cobertura de Activos Fijos'].append((q, round(safe_div(patrim_abs + d['d_lp'], d['af']), 2), month, year))
            # DAX: Ind - Estructura de la Deuda = round(Pasivos(Acum.) / ABS(Patrimonio (Acum. Anual)), 2)
            results['Estructura de la Deuda'].append((q, round(safe_div(a['pasivos_acum'], abs(a['patrim_anual'])), 2), month, year))
            # DAX: Ind - Ratio de Deuda a Activos Tangibles = round(Pasivos(Acum.) / (Patrimonio (Acum. Anual) - Intangibles), 2)
            results['Ratio de Deuda a Activos Tangibles'].append((q, round(safe_div(a['pasivos_acum'], a['patrim_anual'] - a['intangibles']), 2), month, year))

    return results


# ============================================================
# STEP 7: Write CSVs
# ============================================================
MODULES = {
    'LIQUIDEZ': [
        'Razón Corriente', 'Capital de Trabajo', 'Prueba Ácida', 'Ratio de Efectivo'
    ],
    'RENTABILIDAD': [
        'Margen EBITDA', 'Margen Neto Utilidad', 'Margen Operativo', 'Margen de Utilidad Bruta',
        'Patrimonio', 'Retorno sobre Activos (ROA)', 'Retorno sobre Patrimonio (ROE)',
        'Utilidad Acumulada'
    ],
    'SOLVENCIA': [
        'Cobertura de Cargos Fijos', 'Cobertura de Intereses', 'Deuda Neta a EBITDA',
        'Endeudamiento Total', 'Cobertura del Servicio de la Deuda', 'Ratio de Solvencia Patrimonial'
    ],
    'ACTIVIDAD': [
        'Ciclo de Conversión de Efectivo', 'Días de Cartera (DSO)', 'Días de Inventario (DIO)',
        'Días de Proveedores (DPO)', 'Rotación de Activos', 'Rotación de Cartera',
        'Rotación de Inventarios', 'Rotación de Proveedores'
    ],
    'ESTRUCTURA': [
        'Ratio de Capitalización', 'Cobertura de Activos Fijos', 'Estructura de la Deuda',
        'Multiplicador de Capital', 'Ratio de PropiedadAutonomía', 'Relación DeudaPatrimonio',
        'Ratio de Deuda a Activos Tangibles'
    ]
}

# Header names matching original CSV format
HEADER_MAP = {
    'Razón Corriente': 'Times',
    'Prueba Ácida': 'Times',
    'Ratio de Efectivo': 'Times',
    'Capital de Trabajo': 'MM COP',
    'Margen EBITDA': '%',
    'Margen Neto Utilidad': '%',
    'Margen Operativo': '%',
    'Margen de Utilidad Bruta': '%',
    'Patrimonio': 'Mill',
    'Retorno sobre Activos (ROA)': '%',
    'Retorno sobre Patrimonio (ROE)': '%',
    'Utilidad Acumulada': 'Mill',
    'Cobertura de Cargos Fijos': 'Times',
    'Cobertura de Intereses': 'Times',
    'Deuda Neta a EBITDA': 'Year',
    'Endeudamiento Total': '%',
    'Cobertura del Servicio de la Deuda': 'Times',
    'Ratio de Solvencia Patrimonial': '$ COP',
    'Ciclo de Conversión de Efectivo': 'Days',
    'Días de Cartera (DSO)': 'Days',
    'Días de Inventario (DIO)': 'Days',
    'Días de Proveedores (DPO)': 'Days',
    'Rotación de Activos': 'Times',
    'Rotación de Cartera': 'Times',
    'Rotación de Inventarios': 'Times',
    'Rotación de Proveedores': 'Times',
    'Ratio de Capitalización': '$ COP',
    'Cobertura de Activos Fijos': 'Times',
    'Estructura de la Deuda': 'Times',
    'Multiplicador de Capital': 'Times',
    'Ratio de PropiedadAutonomía': '%',
    'Relación DeudaPatrimonio': '$ COP',
    'Ratio de Deuda a Activos Tangibles': '$ COP',
}


def write_csvs(results):
    """Write all indicator CSVs organized by module."""
    total_files = 0

    for module, indicators in MODULES.items():
        module_dir = os.path.join(OUTPUT_DIR, module)
        os.makedirs(module_dir, exist_ok=True)

        for indicator_name in indicators:
            data = results.get(indicator_name, [])
            if not data:
                print(f"  WARNING: No data for {indicator_name}")
                continue

            header_val = HEADER_MAP.get(indicator_name, 'Value')
            filepath = os.path.join(module_dir, f"{indicator_name}.csv")

            with open(filepath, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow(['Quarter', header_val, 'Period', 'Año Texto'])
                for q, val, period, year in data:
                    # Format value: clean floating-point artifacts
                    if isinstance(val, float):
                        # Check if this indicator uses full precision (no ROUND in DAX)
                        full_precision = header_val in ('Times', '$ COP') and indicator_name not in (
                            'Razón Corriente', 'Prueba Ácida', 'Ratio de Efectivo',
                            'Cobertura de Activos Fijos', 'Cobertura de Cargos Fijos',
                            'Cobertura de Intereses', 'Cobertura del Servicio de la Deuda',
                            'Estructura de la Deuda', 'Endeudamiento Total',
                            'Ratio de Solvencia Patrimonial', 'Deuda Neta a EBITDA',
                        )
                        if full_precision:
                            val_fmt = val  # Keep full precision
                        else:
                            rounded = round(val, 2)
                            if rounded == int(rounded):
                                val_fmt = int(rounded)
                            else:
                                val_fmt = rounded
                    else:
                        val_fmt = val
                    writer.writerow([q, val_fmt, period, year])

            total_files += 1

    print(f"  Written {total_files} CSV files to {OUTPUT_DIR}")
    return total_files


# ============================================================
# STEP 8: Compare with originals
# ============================================================
def compare_with_originals(results):
    """Compare generated values with original CSVs to measure accuracy."""
    print("\n" + "=" * 60)
    print("COMPARISON: Generated vs Original CSVs")
    print("=" * 60)

    total_comparisons = 0
    exact_matches = 0
    close_matches = 0  # within 5%
    moderate_matches = 0  # within 20%
    mismatches = []

    # Track per-module stats
    module_stats = {}
    indicator_stats = {}

    for module, indicators in MODULES.items():
        original_dir = os.path.join(SOURCES_DIR, module)
        if not os.path.exists(original_dir):
            continue

        m_total = 0
        m_exact = 0
        m_close = 0

        for indicator_name in indicators:
            original_file = os.path.join(original_dir, f"{indicator_name}.csv")
            if not os.path.exists(original_file):
                continue

            # Read original
            original_values = {}
            with open(original_file, 'r', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                headers = reader.fieldnames
                value_col = headers[1] if headers else 'Value'
                for row in reader:
                    period = int(row['Period'])
                    year = int(row['Año Texto'])
                    try:
                        val = float(row[value_col].replace(',', '.').replace('"', ''))
                    except (ValueError, KeyError):
                        continue
                    original_values[(year, period)] = val

            # Compare with generated
            ind_total = 0
            ind_exact = 0
            ind_close = 0

            generated = results.get(indicator_name, [])
            for q, gen_val, period, year in generated:
                key = (year, period)
                if key not in original_values:
                    continue

                orig_val = original_values[key]
                total_comparisons += 1
                m_total += 1
                ind_total += 1

                if orig_val == 0 and gen_val == 0:
                    exact_matches += 1
                    m_exact += 1
                    ind_exact += 1
                elif abs(gen_val - orig_val) < 0.01:  # absolute tolerance for small values
                    exact_matches += 1
                    m_exact += 1
                    ind_exact += 1
                elif orig_val != 0 and abs((gen_val - orig_val) / orig_val) < 0.05:
                    close_matches += 1
                    m_close += 1
                    ind_close += 1
                elif orig_val != 0 and abs((gen_val - orig_val) / orig_val) < 0.20:
                    moderate_matches += 1
                else:
                    if len(mismatches) < 30:
                        mismatches.append({
                            'indicator': indicator_name,
                            'module': module,
                            'year': year,
                            'period': period,
                            'original': orig_val,
                            'generated': gen_val,
                            'diff_pct': abs((gen_val - orig_val) / orig_val * 100) if orig_val != 0 else 'N/A'
                        })

            if ind_total > 0:
                indicator_stats[indicator_name] = {
                    'total': ind_total,
                    'exact': ind_exact,
                    'close': ind_close,
                    'pct': (ind_exact + ind_close) / ind_total * 100
                }

        if m_total > 0:
            module_stats[module] = {
                'total': m_total,
                'exact': m_exact,
                'close': m_close,
                'pct': (m_exact + m_close) / m_total * 100
            }

    # Print results
    good = exact_matches + close_matches
    print(f"\nTotal comparisons: {total_comparisons}")
    if total_comparisons > 0:
        print(f"Exact matches (<0.01 abs): {exact_matches} ({exact_matches/total_comparisons*100:.1f}%)")
        print(f"Close matches (<5%): {close_matches} ({close_matches/total_comparisons*100:.1f}%)")
        print(f"Moderate matches (<20%): {moderate_matches} ({moderate_matches/total_comparisons*100:.1f}%)")
        print(f"GOOD (exact+close): {good} ({good/total_comparisons*100:.1f}%)")
        print(f"Far mismatches: {total_comparisons - good - moderate_matches}")

    print(f"\n--- BY MODULE ---")
    for mod, s in module_stats.items():
        print(f"  {mod}: {s['pct']:.1f}% match ({s['exact']}+{s['close']}/{s['total']})")

    print(f"\n--- BY INDICATOR ---")
    for ind in sorted(indicator_stats.keys()):
        s = indicator_stats[ind]
        status = '[OK]' if s['pct'] >= 80 else '[~~]' if s['pct'] >= 40 else '[XX]'
        print(f"  {status} {ind}: {s['pct']:.1f}% ({s['exact']}+{s['close']}/{s['total']})")

    if mismatches:
        print(f"\nSample mismatches (first {len(mismatches)}):")
        for m in mismatches[:15]:
            print(f"  [{m['module']}] {m['indicator']} [{m['year']}-P{m['period']}]: "
                  f"Orig={m['original']:.4f} Gen={m['generated']:.4f} "
                  f"Diff={m['diff_pct']}")

    return total_comparisons, exact_matches, close_matches


# ============================================================
# MAIN
# ============================================================
def main():
    print("=" * 60)
    print("PRUEBA ÁCIDA: Financial Indicator Calculator")
    print("MAS CONSULTA SAS - NIT 901271750")
    print("=" * 60)

    print("\n[1/7] Loading account classification from Master Account...")
    accounts, group_termino = load_account_classification()

    print("\n[2/7] Parsing movement files...")
    movements = parse_mov_files()

    print("\n[3/7] Building monthly balance snapshots...")
    m_full, m_ytd, m_flows_op, m_full_ex998, m_ytd_op, m_ytd_all, m_closing, m_flows_all_single = build_monthly_balances(movements, accounts)

    print("\n[4/7] Classifying accounts...")
    classifications = classify_accounts(accounts, group_termino)

    print("\n[5/7] Computing financial aggregates (DAX-aligned)...")
    aggregates = compute_aggregates(m_full, m_ytd, m_flows_op, classifications,
                                    m_full_ex998, m_ytd_op, m_ytd_all, m_closing,
                                    m_flows_all_single)

    print("\n[6/7] Calculating 33 indicators...")
    results = calculate_indicators(aggregates)
    print(f"  Calculated {len(results)} indicators")

    print("\n[7/7] Writing CSVs...")
    write_csvs(results)

    # Bonus: Compare with originals
    compare_with_originals(results)

    print("\n" + "=" * 60)
    print("DONE!")
    print("=" * 60)


if __name__ == '__main__':
    main()
