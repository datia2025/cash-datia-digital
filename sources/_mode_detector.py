"""
MODE DETECTOR: For each indicator, compute both FLOW and CUMULATIVE values, 
then compare both against the original CSV to determine which mode matches best.
"""
import csv, os
from datetime import datetime
from collections import defaultdict
from openpyxl import load_workbook

SOURCES_DIR = os.path.dirname(os.path.abspath(__file__))
MASTER_FILE = os.path.join(SOURCES_DIR, "Master Account.xlsx")
MOV_FILES = [os.path.join(SOURCES_DIR, f"Mov {y}.csv") for y in [2021, 2022, 2023, 2024, 2025]]
ANALYSIS_YEARS = [2023, 2024, 2025]

# Quick load same as main script
def load_data():
    wb = load_workbook(MASTER_FILE, read_only=True, data_only=True)
    ws_list = wb['List']
    group_termino = {}
    for row in ws_list.iter_rows(min_row=2, values_only=True):
        if row and len(row) >= 8:
            grupo = str(row[5]) if row[5] else None
            termino = str(row[7]) if row[7] else None
            if grupo and termino: group_termino[grupo] = termino

    ws_accounts = wb['Accounts']
    accounts = {}
    for row in ws_accounts.iter_rows(min_row=2, values_only=True):
        if not row or not row[1]: continue
        codigo = str(row[1]).strip()
        grupo = str(row[13]) if row[13] else ""
        termino = group_termino.get(grupo, "")
        accounts[codigo] = {'grupo': grupo, 'termino': termino}
    wb.close()
    return accounts, group_termino

def classify(accounts):
    c = {k: set() for k in [
        'activo_corriente', 'activo_no_corriente', 'pasivo_corriente',
        'pasivo_no_corriente', 'patrimonio', 'ingresos', 'gastos', 'costos',
        'cxc', 'inventarios', 'activos_fijos', 'cxp_proveedores', 'cxp_otros',
        'deuda_cp', 'deuda_lp', 'efectivo', 'gastos_financieros',
        'depreciacion', 'amortizacion'
    ]}
    for codigo, info in accounts.items():
        grupo = info.get('grupo', '')
        termino = info.get('termino', '')
        fd = codigo[0] if codigo else '0'
        if fd == '1':
            if termino == 'Largo Termino' or grupo in ('15','16','17'):
                c['activo_no_corriente'].add(codigo)
            else:
                c['activo_corriente'].add(codigo)
            if grupo == '11': c['efectivo'].add(codigo)
            if grupo == '13': c['cxc'].add(codigo)
            if grupo == '14': c['inventarios'].add(codigo)
            if grupo in ('15','16','17'): c['activos_fijos'].add(codigo)
        elif fd == '2':
            if termino == 'Largo Termino':
                c['pasivo_no_corriente'].add(codigo); c['deuda_lp'].add(codigo)
            else:
                c['pasivo_corriente'].add(codigo); c['deuda_cp'].add(codigo)
            if grupo == '22': c['cxp_proveedores'].add(codigo)
            if grupo == '23': c['cxp_otros'].add(codigo)
        elif fd == '3': c['patrimonio'].add(codigo)
        elif fd == '4': c['ingresos'].add(codigo)
        elif fd == '5':
            c['gastos'].add(codigo)
            if codigo.startswith('5305'): c['gastos_financieros'].add(codigo)
            if codigo.startswith('5160') or codigo.startswith('5260'): c['depreciacion'].add(codigo)
            if codigo.startswith('5165') or codigo.startswith('5265'): c['amortizacion'].add(codigo)
        elif fd in ('6','7'): c['costos'].add(codigo)
    return c

def parse_movements():
    movements = defaultdict(lambda: {'debito': 0.0, 'credito': 0.0})
    for filepath in MOV_FILES:
        if not os.path.exists(filepath): continue
        with open(filepath, 'r', encoding='utf-8', errors='replace') as f:
            lines = f.readlines()
        header_idx = None
        for i, line in enumerate(lines):
            if line.startswith('Secuencia,'): header_idx = i; break
        if header_idx is None: continue
        for i in range(header_idx+1, len(lines)):
            line = lines[i].strip()
            if not line or not line[0].isdigit(): continue
            try:
                reader = csv.reader([line])
                fields = next(reader)
                if len(fields) < 12: continue
                desc = fields[7].strip() if len(fields) > 7 else ""
                det = fields[8].strip() if len(fields) > 8 else ""
                if 'cierre anual' in desc.lower() or 'cierre anual' in det.lower(): continue
                fecha = datetime.strptime(fields[1].strip(), '%d/%m/%Y')
                deb = float(fields[10].strip().replace(',','')) if fields[10].strip() else 0
                cre = float(fields[11].strip().replace(',','')) if fields[11].strip() else 0
                movements[(fecha.year, fecha.month, fields[2].strip())]['debito'] += deb
                movements[(fecha.year, fecha.month, fields[2].strip())]['credito'] += cre
            except: continue
    return movements

def build_balances(movements, accounts):
    all_periods = sorted(set((y, m) for y, m, _ in movements.keys()))
    running = defaultdict(float)
    monthly_balances = {}
    monthly_flows = {}
    for year, month in all_periods:
        month_flows = defaultdict(float)
        for (y, m, codigo), amounts in movements.items():
            if y != year or m != month: continue
            fd = codigo[0] if codigo else '0'
            if fd in ('1','5','6','7'):
                delta = amounts['debito'] - amounts['credito']
            elif fd in ('2','3','4'):
                delta = amounts['credito'] - amounts['debito']
            else:
                delta = amounts['debito'] - amounts['credito']
            running[codigo] += delta
            month_flows[codigo] += delta
        monthly_balances[(year, month)] = dict(running)
        monthly_flows[(year, month)] = dict(month_flows)
    return monthly_balances, monthly_flows

def safe_div(a, b):
    if b == 0: return 0
    return a / b

def read_original(module, indicator):
    orig_file = os.path.join(SOURCES_DIR, module, indicator + '.csv')
    if not os.path.exists(orig_file): return None
    vals = {}
    with open(orig_file, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        headers = reader.fieldnames
        val_col = headers[1]
        for row in reader:
            try:
                key = (int(row[u'A\u00f1o Texto']), int(row['Period']))
                vals[key] = float(row[val_col].replace(',','.').replace('"',''))
            except: continue
    return vals

def match_score(generated, original):
    """Compute match % between generated dict and original dict."""
    if not original: return 0, 0
    total = exact = close = 0
    for key, orig in original.items():
        gen = generated.get(key)
        if gen is None: continue
        total += 1
        if orig == 0 and gen == 0:
            exact += 1
        elif abs(gen - orig) < 0.01:
            exact += 1
        elif orig != 0 and abs((gen - orig) / orig) < 0.05:
            close += 1
    return (exact + close), total

def main():
    print("Loading data...")
    accounts, gt = load_data()
    cl = classify(accounts)
    movements = parse_movements()
    mb, mf = build_balances(movements, accounts)
    
    sorted_keys = sorted(mb.keys())
    
    # For each indicator, test both FLOW and CUMULATIVE
    TESTS = {
        # Liquidez
        'Raz\u00f3n Corriente': {
            'module': 'LIQUIDEZ',
            'flow': lambda a, p: safe_div(a['f_ac'], a['f_pc']),
            'cumul': lambda a, p: safe_div(a['c_ac'], a['c_pc']),
        },
        'Capital de Trabajo': {
            'module': 'LIQUIDEZ',
            'flow': lambda a, p: a['f_ac'] - a['f_pc'],
            'cumul': lambda a, p: a['c_ac'] - a['c_pc'],
        },
        'Prueba \u00c1cida': {
            'module': 'LIQUIDEZ',
            'flow': lambda a, p: safe_div(a['f_ac'] - a['f_inv'], a['f_pc']),
            'cumul': lambda a, p: safe_div(a['c_ac'] - a['c_inv'], a['c_pc']),
        },
        'Ratio de Efectivo': {
            'module': 'LIQUIDEZ',
            'flow': lambda a, p: safe_div(a['f_ef'], a['f_pc']),
            'cumul': lambda a, p: safe_div(a['c_ef'], a['c_pc']),
        },
        'Patrimonio': {
            'module': 'RENTABILIDAD',
            'flow': lambda a, p: a['f_pat'],
            'cumul': lambda a, p: a['c_pat'],
        },
        'Endeudamiento Total': {
            'module': 'SOLVENCIA',
            'flow': lambda a, p: safe_div(a['f_pt'], a['f_at']),
            'cumul': lambda a, p: safe_div(a['c_pt'], a['c_at']),
        },
        'Relaci\u00f3n DeudaPatrimonio': {
            'module': 'SOLVENCIA',
            'flow': lambda a, p: safe_div(a['f_pt'], a['f_pat']) if a['f_pat'] != 0 else 0,
            'cumul': lambda a, p: safe_div(a['c_pt'], a['c_pat']) if a['c_pat'] != 0 else 0,
        },
        'Multiplicador de Capital': {
            'module': 'SOLVENCIA',
            'flow': lambda a, p: safe_div(a['f_at'], a['f_pat']) if a['f_pat'] != 0 else 0,
            'cumul': lambda a, p: safe_div(a['c_at'], a['c_pat']) if a['c_pat'] != 0 else 0,
        },
        'Propiedad y Autonom\u00eda': {
            'module': 'SOLVENCIA',
            'flow': lambda a, p: safe_div(a['f_pat'], a['f_at']),
            'cumul': lambda a, p: safe_div(a['c_pat'], a['c_at']),
        },
        'Cobertura de Activos Fijos': {
            'module': 'ESTRUCTURA',
            'flow': lambda a, p: safe_div(a['f_pat'] + a['f_pnc'], a['f_af']) if a['f_af'] != 0 else 0,
            'cumul': lambda a, p: safe_div(a['c_pat'] + a['c_pnc'], a['c_af']) if a['c_af'] != 0 else 0,
        },
        'Estructura de Deuda': {
            'module': 'ESTRUCTURA',
            'flow': lambda a, p: safe_div(a['f_pt'], a['f_pat']) if a['f_pat'] != 0 else 0,
            'cumul': lambda a, p: safe_div(a['c_pt'], a['c_pat']) if a['c_pat'] != 0 else 0,
        },
        'Solvencia Patrimonial': {
            'module': 'ESTRUCTURA',
            'flow': lambda a, p: safe_div(a['f_pt'], a['f_pat']) if a['f_pat'] != 0 else 0,
            'cumul': lambda a, p: safe_div(a['c_pt'], a['c_pat']) if a['c_pat'] != 0 else 0,
        },
        'Relaci\u00f3n Deuda Tangibles': {
            'module': 'ESTRUCTURA',
            'flow': lambda a, p: safe_div(a['f_pt'], a['f_at']) if a['f_at'] != 0 else 0,
            'cumul': lambda a, p: safe_div(a['c_pt'], a['c_at']) if a['c_at'] != 0 else 0,
        },
        'Retorno sobre Activos (ROA)': {
            'module': 'RENTABILIDAD',
            'flow': lambda a, p: safe_div(a['unet'], a['f_at']),
            'cumul': lambda a, p: safe_div(a['unet'], a['c_at']),
        },
        'Retorno sobre Patrimonio (ROE)': {
            'module': 'RENTABILIDAD',
            'flow': lambda a, p: safe_div(a['unet'], a['f_pat']),
            'cumul': lambda a, p: safe_div(a['unet'], a['c_pat']),
        },
    }
    
    # Build aggregates for both modes
    all_aggs = {}
    for (year, month), balances in sorted(mb.items()):
        flows = mf.get((year, month), {})
        def s_a(s): return sum(balances.get(c, 0) for c in s)
        def s_f(s): return sum(flows.get(c, 0) for c in s)
        
        ing = s_f(cl['ingresos'])
        gas = s_f(cl['gastos'])
        cos = s_f(cl['costos'])
        
        all_aggs[(year, month)] = {
            'f_ac': s_f(cl['activo_corriente']),
            'f_anc': s_f(cl['activo_no_corriente']),
            'f_at': s_f(cl['activo_corriente']) + s_f(cl['activo_no_corriente']),
            'f_pc': s_f(cl['pasivo_corriente']),
            'f_pnc': s_f(cl['pasivo_no_corriente']),
            'f_pt': s_f(cl['pasivo_corriente']) + s_f(cl['pasivo_no_corriente']),
            'f_pat': s_f(cl['patrimonio']),
            'f_ef': s_f(cl['efectivo']),
            'f_inv': s_f(cl['inventarios']),
            'f_af': s_f(cl['activos_fijos']),
            'c_ac': s_a(cl['activo_corriente']),
            'c_anc': s_a(cl['activo_no_corriente']),
            'c_at': s_a(cl['activo_corriente']) + s_a(cl['activo_no_corriente']),
            'c_pc': s_a(cl['pasivo_corriente']),
            'c_pnc': s_a(cl['pasivo_no_corriente']),
            'c_pt': s_a(cl['pasivo_corriente']) + s_a(cl['pasivo_no_corriente']),
            'c_pat': s_a(cl['patrimonio']),
            'c_ef': s_a(cl['efectivo']),
            'c_inv': s_a(cl['inventarios']),
            'c_af': s_a(cl['activos_fijos']),
            'unet': ing - gas - cos,
        }
    
    print(f"\n{'='*70}")
    print(f"{'Indicator':<35} {'Flow Match':<15} {'Cumul Match':<15} {'Winner'}")
    print(f"{'='*70}")
    
    for indicator, test in TESTS.items():
        original = read_original(test['module'], indicator)
        if not original:
            print(f"  {indicator:<35} SKIP (no original)")
            continue
        
        flow_vals = {}
        cumul_vals = {}
        for year in ANALYSIS_YEARS:
            for month in range(1, 13):
                key = (year, month)
                if key not in all_aggs: continue
                a = all_aggs[key]
                try:
                    flow_vals[key] = test['flow'](a, None)
                except: flow_vals[key] = 0
                try:
                    cumul_vals[key] = test['cumul'](a, None)
                except: cumul_vals[key] = 0
        
        f_good, f_total = match_score(flow_vals, original)
        c_good, c_total = match_score(cumul_vals, original)
        
        f_pct = f"{f_good}/{f_total} ({f_good/f_total*100:.0f}%)" if f_total > 0 else "N/A"
        c_pct = f"{c_good}/{c_total} ({c_good/c_total*100:.0f}%)" if c_total > 0 else "N/A"
        
        if f_total > 0 and c_total > 0:
            winner = "FLOW" if f_good > c_good else "CUMUL" if c_good > f_good else "TIE"
        else:
            winner = "?"
        
        print(f"  {indicator:<35} {f_pct:<15} {c_pct:<15} {winner}")
    
    print(f"\n{'='*70}")
    print("DONE")

if __name__ == '__main__':
    main()
