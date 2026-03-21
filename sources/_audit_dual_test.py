from openpyxl import load_workbook
import csv
import os
from collections import defaultdict
from datetime import datetime

# Paths
BASE_DIR = r"c:\Users\gutie\OneDrive\FERFILES\DATIA\PROPUESTAS\COLOMBIA\+CONSULTA\LiquidityDashboard\sources"
MASTER_FILE = os.path.join(BASE_DIR, "Master Account.xlsx")
MOV_FILES = [os.path.join(BASE_DIR, f"Mov {y}.csv") for y in [2021, 2022, 2023, 2024, 2025]]
ORIGINAL_CSV = os.path.join(BASE_DIR, "LIQUIDEZ", "Razón Corriente.csv")

def audit_dual_logic():
    # 1. Load Ground Truth
    ground_truth = {}
    if os.path.exists(ORIGINAL_CSV):
        with open(ORIGINAL_CSV, "r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                y = int(row["Año Texto"])
                p = int(row["Period"])
                val = float(row["Times"])
                ground_truth[(y, p)] = val

    # 2. Load Classifications
    wb = load_workbook(MASTER_FILE, read_only=True, data_only=True)
    ws_accounts = wb['Accounts']
    ws_list = wb['List']
    group_termino = {str(row[5]): str(row[7]) for row in ws_list.iter_rows(min_row=2, values_only=True) if row and len(row) >= 8}
    
    ac_set, pc_set = set(), set()
    for row in ws_accounts.iter_rows(min_row=2, values_only=True):
        if not row or not row[1]: continue
        c, g = str(row[1]).strip(), str(row[13])
        t = group_termino.get(g, "")
        if c.startswith('1') and (t == 'Corto Termino' or g in ('11','12','13','14','18','19')): ac_set.add(c)
        if c.startswith('2') and (t == 'Corto Termino' or g in ('22','23','24')): pc_set.add(c)
    
    # 3. Process All Movements (2021-2025)
    movements = defaultdict(lambda: defaultdict(float))
    for mov_path in sorted(MOV_FILES):
        if not os.path.exists(mov_path): continue
        with open(mov_path, 'r', encoding='utf-8', errors='replace') as f:
            reader = csv.reader(f)
            for row in reader:
                if not row or len(row) < 12 or not row[0].isdigit(): continue
                try:
                    dt = datetime.strptime(row[1].strip(), '%d/%m/%Y')
                    movements[(dt.year, dt.month)][row[2].strip()] += (float(row[10].replace(',','')) if row[10] else 0.0) - (float(row[11].replace(',','')) if row[11] else 0.0)
                except: continue

    # 4. Comparative Run
    print(f"{'Periodo':<10} | {'YTD Mode':<10} | {'Full Mode':<10} | {'Original':<10} | {'Result'}")
    print("-" * 75)
    
    full_ac, full_pc_neto = 0, 0
    for y in [2023, 2024, 2025]:
        ytd_ac, ytd_pc_neto = 0, 0
        for m in range(1, 13):
            moves = movements.get((y, m), {})
            ma = sum(moves.get(c, 0) for c in ac_set)
            mp = sum(moves.get(c, 0) for c in pc_set)
            
            ytd_ac += ma; ytd_pc_neto += mp
            full_ac += ma; full_pc_neto += mp # This is simplified, needs history from 2021-22
            
            if y == 2023 and m == 1:
                print(f"AUDIT DEBUG 2023-1: ytd_ac={ytd_ac:.2f} ytd_pc_neto={ytd_pc_neto:.2f} ratio={ytd_ac / (ytd_pc_neto * -1):.4f}")
            
            ytd_val = ytd_ac / (ytd_pc_neto * -1) if ytd_pc_neto != 0 else 0
            # For Full Mode, we need to correctly initialize history
            # (In this run, full_ac/full_pc_neto already accumulated 2021-2022 if processed)
            full_val = full_ac / (full_pc_neto * -1) if full_pc_neto != 0 else 0
            
            orig = ground_truth.get((y, m), 0)
            
            res = "MATCH YTD" if abs(ytd_val - orig) < 0.001 else ("MATCH FULL" if abs(full_val - orig) < 0.001 else "MISMATCH")
            print(f"{y}-P{m:<2} | {ytd_val:10.4f} | {full_val:10.4f} | {orig:10.4f} | {res}")

if __name__ == "__main__":
    # Pre-process 2021-2022 into full_ac/full_pc_neto logic
    audit_dual_logic()
